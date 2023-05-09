from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Q
from Investigations.models import Server, Group, Error, Exceptions, Environment, AdaptorsAndApps, Query, Database, Resource, \
    Configuration, SoapRequest, SoapResponseStatus, Dashboard, ExceptionsCounters, GroupLoggingLevel,\
    QueriesMonitor, SoapRequestHeaders, RepresentationalQuery, SeverityLookup, Service
from .forms import ServerForm, GroupForm, ServerPureForm
import paramiko
#from scp import SCPClient
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import Group as AdminGroup, Permission, User
from django.contrib.contenttypes.models import ContentType
import pysftp
import os.path
from os import listdir
from stat import S_ISDIR, S_ISREG
import subprocess
import mmap
from django.template.response import TemplateResponse
import queue
import threading
import time
import datetime
import requests
from requests.exceptions import HTTPError
from operator import itemgetter
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.executors.pool import ThreadPoolExecutor, ProcessPoolExecutor
from apscheduler import schedulers
import traceback
import json
from openpyxl import Workbook
from openpyxl import load_workbook

import pythoncom
from smb.SMBConnection import SMBConnection
from smb import smb_structs
import wmi
#from pywinos import WinOSClient
import re

import cx_Oracle

from .InfraUtilities import LdapConnection, send_mail
from .MyFawrySearchTools import get_ldap_users, form_ldap_entries, form_table_entries, form_mcc_entries, form_customer_data_tables
from .MonitoringUtilities import addExceptionMonitorRecord, addRequestMonitorRecord, displayExceptionsGraphs, get_query_result, displayQueriesGraphs, getWebsphereQueueinfo, displayQueuesGraphs
from .OperationalTasks import ACH_accounts_job_checker, ACH_success_accounts_job_checker
#from suds.client import Client

environment_nature = Configuration.objects.get(key='env_nature').value
database_used_type = Configuration.objects.get(key='database_used_type').value
database_mongodb = Configuration.objects.get(key='database_mongodb').value
remote_win_out_path = Configuration.objects.get(key='remote_win_out_path', env=environment_nature).value
remote_win_out_name = Configuration.objects.get(key='remote_win_out_name', env=environment_nature).value
local_win_out_path = Configuration.objects.get(key='local_win_out_path', env=environment_nature).value
local_win_out_name = Configuration.objects.get(key='local_win_out_name', env=environment_nature).value
remote_aix_deployment_folder = Configuration.objects.get(key='remote_aix_deployment_folder', env=environment_nature).value
remote_win_deployment_folder = Configuration.objects.get(key='remote_win_deployment_folder', env=environment_nature).value
batch_files_path = Configuration.objects.get(key='batch_files_path', env=environment_nature).value
local_win_deployment_folder = Configuration.objects.get(key='local_win_deployment_folder', env=environment_nature).value
username_property = Configuration.objects.get(key='username_property', env=environment_nature).value
password_property = Configuration.objects.get(key='password_property', env=environment_nature).value

cx_Oracle.init_oracle_client(lib_dir= r"C:\oracle\instantclient-basic-windows.x64-21.3.0.0.0\instantclient_21_3")


ldap_params = ['cn', 'sn', 'uid', 'mobile', 'mail', 'roomNumber', 'employeeNumber', 'internationalisdnnumber', 'title',
                'businessCategory']


def get_logged_in_user(request):
    current_user = request.user
    return current_user


@login_required(login_url="/login")
@permission_required('auth.search_myfawry_users', login_url="/login", raise_exception=True)
def userinfopage(request):
    return render(request, 'user_info.html')


@csrf_exempt
@login_required(login_url="/login")
@permission_required('auth.search_myfawry_users', login_url="/login", raise_exception=True)
def GetCustFn(request):
    if request.method == 'POST':
        mobile = request.POST['mobile']
        mail = request.POST['mail']
        print(mail)
        print(mobile)

        if mobile == '' and mail == '':
            return HttpResponse('please enter mail or mobile')
        elif mobile != '' and mail != '':
            return HttpResponse('please select only one entry')
        elif mobile != '':
            mob_reg_ex = '(^01([0-2]|5){1}[0-9]{8}$)|(^(?:\+971|00971|0)?(?:50|51|52|55|56|2|3|4|6|7|9)d{7}$)|(^(009665|9665|\+9665|05|5)(5|0|3|6|4|9|1|8|7)([0-9]{7})$)|(^[3|5|6|7]([0-9]{7})$)|(^[5|6|9]([0-9]{7})$)'
            if not re.match(mob_reg_ex, mobile):
                return HttpResponse('please enter valid mobile no')
        elif mail != '':
            mail_regex = '^[_A-Za-z0-9-\+]+(\.[_A-Za-z0-9-]+)*@[A-Za-z0-9-]+(\.[A-Za-z0-9]+)*(\.[A-Za-z]{2,})$'
            if not re.match(mail_regex, mail):
                return HttpResponse('please enter valid email')

        try:
            ldap_conn = LdapConnection(get_logged_in_user(request))

            in_queue = queue.Queue()
            cifs = []
            mobiles = []

            if mobile != '':
                query = "(mobile=" + mobile + ")"
                result = get_ldap_users(get_logged_in_user(request), query, ldap_conn)
            elif mail != '':
                query = "(mail=" + mail + ")"
                result = get_ldap_users(get_logged_in_user(request), query, ldap_conn)
        except Exception as e:
            print(e.with_traceback())
            return HttpResponse('please check you are connected to vpn or if ldap down')
        finally:
            ldap_conn.unbind()

        if len(result) >= 1:
            print(result)
            table = form_ldap_entries(get_logged_in_user(request), result)
            in_queue.put(table)

            for entry in result:
                try:
                    value = entry[ldap_params[6]]
                    print(type(value), str(value))
                    cif_without_code = str(value)
                    country_code = entry[ldap_params[7]]
                    if country_code == None or country_code == '':
                        cif = cif_without_code
                    else:
                        cif = cif_without_code

                except TypeError:
                    return HttpResponse('One of LDAP entries has no CIF')
                if cif not in cifs:
                    cifs.append(cif)

                try:
                    value = entry[ldap_params[3]]
                    mobile = value[0].strip()
                except TypeError:
                    return HttpResponse('One of LDAP entries has no mobile')
                if mobile not in mobiles:
                    mobiles.append(mobile)
        else:
            return HttpResponse('no results found')
        print(cifs)
        print(mobiles)

        try:
            table += form_customer_data_tables(request, cifs, mobiles)
        except cx_Oracle.Error as error:
            print(error)

        stdout = table
        return HttpResponse(stdout)
    return HttpResponse('fail')


@login_required(login_url="/login")
def index(request):
    return render(request, 'index.html')


@login_required(login_url="/login")
@permission_required('auth.can_check_log_files', login_url="/login", raise_exception=True)
def grepForLogFilesInfoPage(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging')).order_by('groupName')
    servers = Server.objects.all()
    context = {
        'groups': groups,
        'servers': servers,
    }
    print(servers)
    return render(request, 'grep_for_log_files.html', context)


@login_required(login_url="/login")
@permission_required('auth.command_executer', login_url="/login", raise_exception=True)
def executeCommandPage(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging')).order_by('groupName')
    servers = Server.objects.all()
    context = {
        'groups': groups,
        'servers': servers,
    }
    return render(request, 'execute_command.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_handle_obie_check', login_url="/login", raise_exception=True)
def obieHandlingPage(request):
    return render(request, 'obi_handling_page.html')


@login_required(login_url="/login")
@permission_required('auth.can_control_scheduler_tasks', login_url="/login", raise_exception=True)
def SchHandlingPage(request):
    return render(request, 'Sch_handling_page.html')


@login_required(login_url="/login")
@permission_required('auth.can_check_log_requests', login_url="/login", raise_exception=True)
def getMessagesInfoPage(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging',LogsOpened='true').order_by('groupName')
    elif environment_nature in ['local']:
        #groups = Group.objects.all().order_by('groupName')
        groups = Group.objects.filter(LogsOpened='true').order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging'),LogsOpened='true').order_by('groupName')
    servers = Server.objects.all()
    context = {
        'groups': groups,
        'servers': servers,
    }
    return render(request, 'get_messages_info_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_change_log_level', login_url="/login", raise_exception=True)
def changeLogLevel(request):
    result = GroupLoggingLevel.objects.values('group__groupName')
    groups = []
    for entry in result:
        groups.append(entry['group__groupName'])
    print(groups)
    """
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.filter(LogsOpened='false').order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging'),LogsOpened='false').order_by('groupName')
    """
    servers = Server.objects.all()

    context = {
        'groups': groups,
        'servers': servers,
    }
    return render(request, 'change_log_info_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_get_string_count', login_url="/login", raise_exception=True)
def count_exceptions_page(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging')).order_by('groupName')
    servers = Server.objects.all()
    exceptions = Exceptions.objects.all()
    context = {
        'groups': groups,
        'servers': servers,
        'exceptions': exceptions,
    }
    return render(request, 'count_exceptions_info_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_monitor_version_downloads', login_url="/login", raise_exception=True)
def downloadMonitorPage(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging')).order_by('groupName')
    servers = Server.objects.all()
    exceptions = Exceptions.objects.all()
    context = {
        'groups': groups,
        'servers': servers,
        'exceptions': exceptions,
    }
    return render(request, 'Download_monitor.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_check_env_health', login_url="/login", raise_exception=True)
def check_env_health_page(request):
    environments = Environment.objects.all()
    context = {
        'environments': environments,
    }
    return render(request, 'check_env_health_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.check_rep_query_monitor_page', login_url="/login", raise_exception=True)
def check_rep_query_monitor_page(request):
    environments = Environment.objects.all()
    context = {
        'environments': environments,
    }
    return render(request, 'check_rep_queries_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.check_env_health_graphs', login_url="/login", raise_exception=True)
def check_env_health_graphs_page(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging'), allow_monitoring=True).order_by('groupName')
    context = {
        'groups': groups,
    }
    return render(request, 'get_env_health_graphs.html', context)


@login_required(login_url="/login")
@permission_required('auth.check_env_health_graphs', login_url="/login", raise_exception=True)
def check_queues_health_graphs_page(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging'), allow_monitoring=True).order_by('groupName')
    context = {
        'groups': groups,
    }
    return render(request, 'get_queues_health_graphs.html', context)


@login_required(login_url="/login")
@permission_required('auth.check_queries_health_graphs', login_url="/login", raise_exception=True)
def check_queries_health_graphs_page(request):

    databases = Database.objects.filter(~Q(environment__envName='staging'), allow_monitoring=True).order_by('name')
    severities = SeverityLookup.objects.all()
    context = {
        'databases': databases,
        'severities': severities,
    }
    return render(request, 'get_queries_health_graphs.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_check_using_soap', login_url="/login", raise_exception=True)
def check_soap_status_page(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging')).order_by('groupName')
    servers = Server.objects.all()
    apps = AdaptorsAndApps.objects.all()
    context = {
        'groups': groups,
        'servers': servers,
        'apps':apps,
    }
    return render(request, 'check_soap_status_info_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_check_using_soap', login_url="/login", raise_exception=True)
def check_service_using_soap(request):
    environments = Environment.objects.all()

    context = {
        'environments': environments,
    }
    return render(request, 'check_service_soap_info_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_check_archive_requests', login_url="/login", raise_exception=True)
def check_archives_page(request):
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.filter(~Q(archiveCountPerDay='0')).order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging'),~Q(archiveCountPerDay='0')).order_by('groupName')
    servers = Server.objects.all()
    context = {
        'groups': groups,
        'servers': servers,
    }
    return render(request, 'archive_info_page.html', context)


@login_required(login_url="/login")
@permission_required('auth.can_control_deployment_tasks', login_url="/login", raise_exception=True)
def check_deploy_page(request):
    apps = AdaptorsAndApps.objects.all()
    if environment_nature in ['staging']:
        groups = Group.objects.filter(environments__envName='staging').order_by('groupName')
    elif environment_nature in ['local']:
        groups = Group.objects.all().order_by('groupName')
    else:
        groups = Group.objects.filter(~Q(environments__envName='staging')).order_by('groupName')
    servers = Server.objects.all()
    context = {
        'apps': apps,
        'groups': groups,
        'servers': servers,
    }
    return render(request, 'deploypage.html', context)


@login_required(login_url="/login")
def check_dashboard_page(request):
    dashboards = Dashboard.objects.all()
    context = {
        'dashboards': dashboards,
    }
    return render(request, 'dashboardspage.html', context)


class ServerConnection:
    def __init__(self, machine):
        self.machine = machine

    def connect_ssh(self):
        if self.machine.os == 'aix':
            self.client = paramiko.SSHClient()
            self.client.load_system_host_keys()
            self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.client.connect(self.machine.IP, 22, self.machine.app_user, self.machine.app_password)
        if self.machine.os == 'linux':
            self.client = paramiko.SSHClient()
            self.client.load_system_host_keys()
            self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.client.connect(self.machine.IP, 22, self.machine.app_user, self.machine.app_password)
        elif self.machine.os == 'windows':
            pythoncom.CoInitialize()
            self.client = wmi.WMI(self.machine.IP, user=self.machine.app_user, password=self.machine.app_password)

    def disconnect_ssh(self):
        if self.machine.os in ('aix', 'linux'):
            self.client.close()
        elif self.machine.os == 'windows':
            pythoncom.CoUninitialize()

    def exec_command(self, command, skip_print_log=None):
        print(command)
        if self.machine.os in ('aix', 'linux'):
            stdin, stdout, stderr = self.client.exec_command(command)
            temp = stdout.read()
            if skip_print_log is None:
                print(temp.decode("utf-8"))

            """while not self.client.recv_exit_status():
                if self.client.recv_ready():
                    data = self.client.recv(1024)
                    while data:
                        print data
                        data = chan.recv(1024)

                if self.client.recv_stderr_ready():
                    error_buff = self.client.recv_stderr(1024)
                    while error_buff:
                        print error_buff
                        error_buff = self.client.recv_stderr(1024)
                exist_status = self.client.recv_exit_status()
                if 0 == exist_status:"""
            return temp.decode("utf-8")

        elif self.machine.os == 'windows':
            """
            tool = WinOSClient(self.machine.IP, self.machine.app_user, self.machine.app_password, logger_enabled=False)
            response = tool.run_cmd(command='ipconfig > c:\\deployment\\dir.log')
            print(response.ok)
            """
            psexec_path = local_win_deployment_folder + r'\PSTools\psexec.exe'
            cmd = r' \\' + self.machine.IP + r'\deployment\new.bat "' + command + '" ' + remote_win_out_path + '\\' + remote_win_out_name
            print(psexec_path + cmd)
            process = subprocess.Popen(psexec_path + cmd, shell=True, stdout=subprocess.PIPE)
            (outs, errs) = process.communicate(input=None, timeout=None)
            print(outs)

            """
            exec_cmd = r'cmd.exe /c ' + remote_win_out_path + '\\New.bat ' + command + ' ' + remote_win_out_path + '\\' + remote_win_out_name + ''
            print(exec_cmd)
            process_id, result = self.client.Win32_Process.Create(CommandLine=exec_cmd)
            
            watcher = self.client.watch_for(
                notification_type="Deletion",
                wmi_class="Win32_Process",
                delay_secs=1,
                ProcessId=process_id
            )
            watcher()
            """
            #print(result)
            """
            conn = SMBConnection(self.machine.app_user, self.machine.app_password, 'test', self.machine.hostname, use_ntlm_v2=True)
            conn.connect(self.machine.IP, 139)
            with open(local_win_out_name, 'wb+') as fp:
                conn.retrieveFile('deployment', '/' + remote_win_out_name, fp)
            """
            cmnd = 'copy ' + '"\\\\' + self.machine.IP + '\\deployment\\' + remote_win_out_name + '" "' + local_win_out_path + '"'
            print(cmnd)
            stream = os.popen(cmnd)
            output = stream.read()
            print(output)
            with open(local_win_out_path + '\\' +local_win_out_name, 'r') as fr:
                lines = fr.readlines()
                output = ''
                for line in lines:
                    output += line
                print(output)
            #conn.close()
            return output

    def move_remotely_to_win(self, folder_name):
        if self.machine.os in ('aix', 'linux'):
            try:
                process = subprocess.Popen([folder_name, self.machine.IP, username_property, password_property], shell=True, stdout=subprocess.PIPE)
                (outs, errs) = process.communicate(input=None, timeout=None)
                print(outs)
                return outs
            except TimeoutError as e:
                print(str(e))
                process.kill()
                print(process.pid + " on server: " + self.machine.IP + " is killed.")
                return b''
            except Exception as e:
                print(str(e))
                return str(e)
            finally:
                process.terminate()

        elif self.machine.os == 'windows':
            try:
                conn = SMBConnection(self.machine.app_user, self.machine.app_password, 'test', self.machine.hostname,
                                use_ntlm_v2=True)
                conn.connect(self.machine.IP, 139)
                files = conn.listPath('deployment', folder_name, search=65591, pattern='*', timeout=30)
                if len(files) > 2:
                    for x in range(2, len(files)-1):
                        print(str(files[x].filename))
                    conn.deleteFiles('deployment', '/' + folder_name +'/*', delete_matching_folders=True, timeout=30)
            except smb_structs.OperationFailure as e:
                print(str(e))
                conn.createDirectory('deployment', folder_name)
            finally:
                path = local_win_deployment_folder + '\\' + folder_name
                copy_files_remotely(path, folder_name, self.machine.IP, conn)
                conn.close()
            return 'moving resources'.encode("utf-8")


def copy_files_remotely(path, folder_name, IP, conn):
    contents = listdir(path)
    for content in contents:
        if os.path.isfile(os.path.join(path, content)):
            cmnd = 'copy "' + path + '\\' + content + '" "\\\\' + IP + '\\deployment\\' + folder_name + '"'
            print(cmnd)
            stream = os.popen(cmnd)
            output = stream.read()
            print(output)
        else:
            conn.createDirectory('deployment', folder_name + '/' + content)
            copy_files_remotely(os.path.join(path, content), folder_name + '\\' + content, IP, conn)


@csrf_exempt
@permission_required('auth.can_check_log_files', login_url="/login", raise_exception=True)
def search_log_file_fn(request):
    print('=============================================== New Thread in search_log_file_fn =======================================')
    try:
        if request.method == 'POST':
            asynch = request.POST['Async']
            if 'Async' in request.POST:
                if 'search_group' in request.POST:
                    search_group = request.POST['search_group']
                    matchingIPs = Server.objects.filter(group_name__groupName=search_group)
                    trx_date = request.POST['trx_date']
                    trx_hr = request.POST['trx_hour']
                    log_type = request.POST['is_switch']
                    in_queue = queue.Queue()
                    threads_list = list()


                    for s in matchingIPs:

                        t = threading.Thread(target=greptool, args=(s.IP, s.log_path, asynch, in_queue, log_type, trx_date,
                                                                trx_hr, s.logfile_prefix))
                        t.start()
                        threads_list.append(t)

                    for t in threads_list:
                        t.join()

                    stdout = ''
                    while not in_queue.empty():
                        result = in_queue.get()
                        stdout += result
                        stdout += '\n'
                    print(stdout)
                return HttpResponse(stdout)
    except Exception as e:
        return HttpResponse("Error: failed to get files for server %s , exception: %s" % (hostname, str(e)))
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.can_control_deployment_tasks', login_url="/login", raise_exception=True)
def deploy_fn(request):
    try:
        print(
            '=============================================== New Thread in deploy_fn =======================================')
        if request.method == 'POST':
            app_type = request.POST['app_type']
            action_type = request.POST['action_type']
            ip_list_string = request.POST.get('ip_list')
            ip_list = ip_list_string.split(':')
            print(ip_list)
            output = ''

            if app_type in ('promo'):
                print('not normal')
            elif app_type in ('BWH-Mappers', 'BWH-Xip', 'BWH-Portal', 'BWH-Online'):
                print('Deploying BWH')

                threads_list = list()
                in_queue = queue.Queue()

                for ip in ip_list:
                    if ip != '':
                        t = threading.Thread(target=deploy_dot_net, args=(
                            ip, in_queue))
                        t.start()
                        threads_list.append(t)

                for t in threads_list:
                    t.join()

                while not in_queue.empty():
                    result = in_queue.get()
                    output += result
                    output += '\n'

            elif app_type == 'Switch':
                online_bar = request.POST['online_bar']
                offline_bar = request.POST['offline_bar']

                for ip in ip_list:
                    if ip != '':
                        srv = Server.objects.get(IP=ip)
                        output += srv.IP
                        output += "\n====================================\n"
                        #output += (move_to_remote(srv, 'sw_ear')).decode("utf-8")
                        try:
                            output += "**** Deploying the bar ****\n"
                            print("**** Deploying the bar ****")

                            client = paramiko.SSHClient()
                            client.load_system_host_keys()
                            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            client.connect(srv.IP, 22, srv.app_user, srv.app_password)
                            channel = client.get_transport().open_session()

                            command = 'cd /opt/IBM/iib-10.0.0.17/server/bin ; ./mqsideploy -p 4414 -b SwitchNode1 -e switch_eg -a "/var/mqm/deployment/' + online_bar + '" -w 300 -m'
                            print(command)
                            output += command
                            stdout = ''
                            channel.exec_command(command)
                            print(channel.recv_exit_status())

                            while not channel.recv_exit_status():
                                if channel.recv_ready():
                                    data = channel.recv(1024)
                                    while data:
                                        print(data)
                                        data = channel.recv(1024)

                                if channel.recv_stderr_ready():
                                    error_buff = channel.recv_stderr(1024)
                                    while error_buff:
                                        print(error_buff)
                                        error_buff = channel.recv_stderr(1024)
                                exist_status = channel.recv_exit_status()
                                if 0 == exist_status:
                                    break

                            #temp = stdout.read()
                            #print(temp.decode("utf-8"))
                            output += '\n'
                            #output += temp.decode("utf-8")

                            command = 'cd /opt/IBM/iib-10.0.0.17/server/bin ; mqsideploy -p 4414 -b SwitchNode1 -e switch_eg1 -a "/var/mqm/deployment/' + online_bar + '" -w 300 -m'
                            print(command)
                            output += '\n'
                            output += command
                            stdin, stdout, stderr = client.exec_command(command)
                            temp = stdout.read()
                            print(temp.decode("utf-8"))
                            output += '\n'
                            output += temp.decode("utf-8")

                            command = 'cd /opt/IBM/iib-10.0.0.17/server/bin ; mqsideploy -p 4414 -b SwitchNode1 -e switch_eg2 -a "/var/mqm/deployment/' + offline_bar + '" -w 300 -m'
                            print(command)
                            output += '\n'
                            output += command
                            #stdin, stdout, stderr = client.exec_command(command)
                            temp = stdout.read()
                            print(temp.decode("utf-8"))
                            output += '\n'
                            output += temp.decode("utf-8")

                        except EOFError:
                            err = 'cannot connect machine : ' + srv.IP
                            print(err)
                            return err.encode("utf-8")
                        except paramiko.ssh_exception.AuthenticationException:
                            err = 'Invalid username or password for machine : ' + srv.IP
                            print(err)
                            return err.encode("utf-8")
                        finally:
                            client.close()

            else:
                new_ear = request.POST['new_ear']
                old_ear = request.POST['old_ear']
                hasResources = request.POST['hasResources']
                hasQueues = request.POST['hasQueues']
                startEar = request.POST['startEar']

                threads_list = list()
                in_queue = queue.Queue()

                for ip in ip_list:
                    if ip != '':
                        t = threading.Thread(target=deploy_websphere, args=(
                            request, ip, action_type, app_type, new_ear, old_ear, hasResources, hasQueues, startEar, in_queue))
                        t.start()
                        threads_list.append(t)

                for t in threads_list:
                    t.join()

                while not in_queue.empty():
                    result = in_queue.get()
                    output += result
                    output += '\n'

            return HttpResponse(output)

        return HttpResponse("fail")

    except Exception as e:
        return HttpResponse("Exception occured: " + str(e))


def deploy_dot_net(ip, out_queue):
    srv = Server.objects.get(IP=ip)
    output = '\n'
    if srv is not None:
        try:
            output += srv.IP
            output += "\n====================================\n"
            print('**** Moving Resources folders ****')
            output += '**** Moving Resources folders ****\n'
            tmp = move_to_remote(srv, 'resources').decode("utf-8")
            print(tmp)
            output += tmp
            output += '\n'
            print('**** Moving scripts ****')
            print(move_to_remote(srv, 'scripts').decode("utf-8"))

            cmd = r' \\' + srv.IP + ' -u ' + srv.app_user + ' -p ' + srv.app_password + ' -e -s ' + \
                remote_win_deployment_folder + r'\scripts\deploy_windows.bat "' + remote_win_deployment_folder + r'\backup" "' + \
                srv.resources_path + '" "' + remote_win_deployment_folder + r'\resources" ' + srv.bin_path

            output += cmd

            psexec_path = local_win_deployment_folder + r'\PSTools\psexec.exe'
            process = subprocess.Popen(psexec_path + cmd, shell=True, stdout=subprocess.PIPE)
            (outs, errs) = process.communicate(input=None, timeout=None)
            print(outs)
            output += outs
            out_queue.put(output)

        except TimeoutError as e:
            print(str(e))
            output += '\n'
            output += str(e)
            output += '\n'
            process.kill()
            msg = process.pid + " on server: " + srv.IP + " is killed."
            print(msg)
            output += msg
            out_queue.put(output)
        except Exception as e:
            print(str(e))
            output += '\n'
            output += str(e)
            output += '\n'
            out_queue.put(output)
        finally:
            process.terminate()
            return


def deploy_websphere(request, ip, action_type, app_type, new_ear, old_ear, hasResources, hasQueues, startEar, in_queue,):
    srv = Server.objects.get(IP=ip)
    output = '\n'
    if srv is not None:
        output += srv.IP
        output += "\n====================================\n"
        #move_to_remote(srv, 'scripts')

        try:
            if action_type == 'change resources' or action_type == 'deployment':

                resource = Resource.objects.filter(serverIP__IP=srv.IP, appType__Code=app_type)

                if len(resource) == 0:
                    group = Group.objects.get(groupName=srv.group_name)
                    resource = Resource.objects.filter(group__groupName=group, appType__Code=app_type)
                if len(resource) == 0:
                    print('please link resources to this machine : ' + srv.IP)
                    print(resource)
                    in_queue.put('please link resources to this machine : ' + srv.IP)
                resource = resource.first()

                if resource.Code != '':
                    sharedLibs = (resource.Code).split(',')
                resFolders = (resource.srvFolderNames).split(',')


            srv_conn = ServerConnection(srv)
            srv_conn.connect_ssh()

            if action_type == 'deployment':
                print('**** Deploying the Ear ****')
                output += '\n**** Deploying the Ear ****\n'
                #print(move_to_remote(srv, 'ear').decode("utf-8"))

                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/deploy-application.py " + \
                          remote_aix_deployment_folder + "/ear/" + new_ear + " " + new_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin^ -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\deploy-application.py " + \
                              remote_win_deployment_folder + "\ear\\" + new_ear + " " + new_ear + '"'

                output += srv_conn.exec_command(command)

                if resource.Code != '':
                    for sl in sharedLibs:
                        if srv.os in ('aix', 'linux'):
                            command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_aix_deployment_folder + "/scripts/assign-shared-lib.py " + sl + " " + new_ear
                        elif srv.os == 'windows':
                            command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                                  + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                                  + " -f " + remote_win_deployment_folder + "\scripts\\assign-shared-lib.py " + sl + " " + new_ear
                        print("**** Attaching shared library ****")
                        output += '\n**** Attaching shared library ****\n'
                        output += srv_conn.exec_command(command)

                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/disable-auto-start.py " + old_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\disable-auto-start.py " + old_ear
                print('**** Disabling auto start of old ear ****')
                output += '\n**** Disabling auto start of old ear ****\n'
                output += srv_conn.exec_command(command)

            if (hasResources == 'true' and action_type == 'deployment') or action_type == 'change resources':
                print('**** Moving Resources folders ****')
                output += '\n**** Moving Resources folders ****'
                print(move_to_remote(srv, 'resources').decode("utf-8"))
                for folder in resFolders:

                    if srv.os in ('aix', 'linux'):
                        command = 'chmod -R 777 ' + srv.resources_path + ' ; cp -R ' + srv.resources_path + '/' + folder + ' ' + remote_aix_deployment_folder + '/backup/' + folder
                    elif srv.os == 'windows':
                        command = 'copy -R ' + srv.resources_path + '\\' + folder + ' ' + remote_win_deployment_folder + '\\backup\\' + folder
                    output += '\n**** Taking backup of folder : ' + folder + ' ****'
                    print('Taking backup of folder : ' + folder + ' --->')
                    temp = srv_conn.exec_command(command)
                    if temp == '':
                        output += ' success...'
                    else:
                        output += ' fail...'

                for folder in resFolders:
                    if srv.os in ('aix', 'linux'):
                        command = 'cp -R ' + remote_aix_deployment_folder + '/resources/' + folder + ' ' + srv.resources_path
                    elif srv.os == 'windows':
                        command = 'copy -R ' + remote_win_deployment_folder + '\\resources\\' + folder + ' ' + srv.resources_path
                    output += '\n**** Moving Resource folder : ' + folder + ' ****'
                    print('Moving folder : ' + folder + ' --->')
                    temp = srv_conn.exec_command(command)
                    if temp == '':
                        output += ' success...'
                    else:
                        output += ' fail...'

            if action_type == 'move resources':
                print('**** Moving Resources folders ****')
                output += '\n**** Moving Resources folders ****'
                print(move_to_remote(srv, 'resources').decode("utf-8"))

                comd = 'cd ' + remote_aix_deployment_folder + '/resources/ ; ls'
                temp = srv_conn.exec_command(comd)
                temp = temp.split()
                print(temp)
                for folder in temp:
                    if srv.os in ('aix', 'linux'):
                        command = 'cp -R ' + remote_aix_deployment_folder + '/resources/' + folder + ' ' + srv.resources_path
                    elif srv.os == 'windows':
                        command = 'copy -R ' + remote_win_deployment_folder + '\\resources\\' + folder + ' ' + srv.resources_path
                    output += '\n**** Moving Resource ****'
                    temp = srv_conn.exec_command(command)
                    if temp == '':
                        output += ' success...'
                        output += folder
                    else:
                        output += ' fail...'
            if action_type == 'move lib':
                print('**** Moving lib folders ****')
                output += '\n**** Moving lib folders ****'
                print(move_to_remote(srv, 'resources').decode("utf-8"))

                for folder in ['ext', 'ext_SOF', 'jdbc', 'Mohasel_Lib']:
                    if srv.os in ('aix', 'linux'):
                        command = 'cp -R ' + remote_aix_deployment_folder + '/lib/' + folder + ' /opt/IBM/WebSphere/AppServer/lib'
                    elif srv.os == 'windows':
                        command = 'copy -R ' + remote_win_deployment_folder + '\\resources\\' + folder + ' ' + srv.resources_path
                    output += '\n**** Moving Resource ****'
                    temp = srv_conn.exec_command(command)
                    if temp == '':
                        output += ' success...'
                        output += folder
                    else:
                        output += ' fail...'

            if (hasQueues == 'true' and action_type == 'deployment') or action_type == 'deploy queues':
                output += '\n\n**** Deploying Queues ****\n'
                print('**** Deploying Queues ****')
                #print(move_to_remote(srv, 'queues').decode("utf-8"))
                if srv.os in ('aix', 'linux'):
                    command = 'cd ' + remote_aix_deployment_folder + '/queues ; ls -1'
                elif srv.os == 'windows':
                    command = 'cd ' + remote_win_deployment_folder + '\queues ; dir /b'

                temp = srv_conn.exec_command(command)
                output += temp
                output += '\n'
                queuesList = temp.split('\n')
                for qname in queuesList:
                    if qname != '':
                        if srv.os in ('aix', 'linux'):
                            command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                                  + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                                  + " -f " + remote_aix_deployment_folder + "/queues/" + qname
                        elif srv.os == 'windows':
                            command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                                      + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                                      + " -f " + remote_win_deployment_folder + "\queues\\" + qname
                        output += srv_conn.exec_command(command)
                        output += '\n'

            if startEar == 'true' and action_type == 'deployment':
                output += '\n**** Starting Ear ****\n'
                print('**** Starting Ear ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/start-stop-app.py " + old_ear + ' ' + new_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\start-stop-app.py " + old_ear + ' ' + new_ear
                output += srv_conn.exec_command(command)
                output += '\n'

            if action_type == 'stop application':
                output += '**** Stop Ear ****\n'
                print('**** Stop Ear ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/stop-app.py " + old_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\stop-app.py " + old_ear
                temp = srv_conn.exec_command(command)
                resp = 'com.ibm.ws.exception.RuntimeWarning: Application ' + old_ear + ' not started'
                if resp in temp:
                    output += temp
                    #output += 'Application ' + old_ear + ' stopped successfully'

                else:
                    output += temp
                    #output += 'Failed to stop Application ' + old_ear
                output += '\n'

            if action_type == 'start application':
                output += '**** start Ear ****\n'
                print('**** Start Ear ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/start-app.py " + new_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\start-app.py " + new_ear
                temp = srv_conn.exec_command(command)
                resp = 'cuedition=BASE already started'
                if resp in temp:
                    output += temp
                    #output += 'Application ' + old_ear + ' started successfully'
                else:
                    output += temp
                    #output += 'Failed to start Application ' + old_ear
                output += '\n'

            if action_type == 'uninstall application':
                output += '**** uninstall Ear ****\n'
                print('**** uninstall Ear ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/uninstall-application.py " + old_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + r"\scripts\uninstall-application.py " + old_ear
                output += srv_conn.exec_command(command)
                output += '\n'

            if action_type == 'create shared lib':
                resource_name = request.POST['resource_name']
                resource_path = request.POST['resource_path']
                output += '**** creating shared lib ****\n'
                print('**** creating shared lib ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/create-shared-lib.py " + resource_name + ' ' + resource_path
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\create-shared-lib.py " + resource_name + ' ' + resource_path
                output += srv_conn.exec_command(command)
                output += '\n'

            if action_type == 'Running Apps':
                output += '**** Get Running Apps ****\n'
                print('**** Get Running Apps ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/getRunningApps.py"
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\getRunningApps.py"
                output += srv_conn.exec_command(command)
                output += '\n'

            if action_type == 'enable auto start':
                output += '**** enable auto start ****\n'
                print('**** enable auto start ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/enable-auto-start.py " + old_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\enable-auto-start.py " + old_ear
                temp = srv_conn.exec_command(command)
                resp = 'com.ibm.ws.exception.RuntimeWarning: Application ' + old_ear + ' not started'
                if resp in temp:
                    output += temp
                    #output += 'Application ' + old_ear + ' stopped successfully'

                else:
                    output += temp
                    #output += 'Failed to stop Application ' + old_ear
                output += '\n'

            if action_type == 'disable auto start':
                output += '**** disable auto start ****\n'
                print('**** disable auto start ****')
                if srv.os in ('aix', 'linux'):
                    command = "cd " + srv.bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " \
                          + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                          + " -f " + remote_aix_deployment_folder + "/scripts/disable-auto-start.py " + old_ear
                elif srv.os == 'windows':
                    command = '"' + srv.bin_path + '\wsadmin" -lang jython -conntype SOAP -port ' \
                              + srv.soap_port + " -username " + srv.app_user + " -password " + srv.soap_pass \
                              + " -f " + remote_win_deployment_folder + "\scripts\disable-auto-start.py " + old_ear
                temp = srv_conn.exec_command(command)
                resp = 'com.ibm.ws.exception.RuntimeWarning: Application ' + old_ear + ' not started'
                if resp in temp:
                    output += temp
                    #output += 'Application ' + old_ear + ' stopped successfully'

                else:
                    output += temp
                    #output += 'Failed to stop Application ' + old_ear
                output += '\n'

            if action_type == 'Restart Websphere App Server' and str(srv.middleware_tech) == 'Websphere':
                command = 'ps -aef | grep WebSphere'
                output += '\n**** Checking running processes ****\n'
                print('**** Checking running processes ****')
                temp_out = srv_conn.exec_command(command)
                print(temp_out)
                output += temp_out
                while temp_out.find('  ') != -1:
                    temp_out = temp_out.replace('  ', ' ')
                temp_out = temp_out.split('\n')

                result = []
                for line in temp_out:
                    line_out = line.split(' ')
                    result.append(line_out)
                print(result)
                output += '\n'
                was_processes = []
                for i in range(0, len(result)):
                    for valu in result[i]:
                        if valu == 'com.ibm.ws.runtime.WsServer':
                            if valu[-1] == 'dmgr':
                                continue
                            was_processes.append(result[i])
                print(was_processes)

                output += '\n**** Killing WAS processes ****\n'
                print('**** Killing WAS processes ****')

                for i in range(0, len(was_processes)):
                    command = 'kill -9 ' + was_processes[i][1]
                    output += command
                    output += '\n'
                    output += srv_conn.exec_command(command)

                print("#checking if theres any WAS processes up before deleting temp folders")
                command = 'ps -aef | grep WebSphere'
                temp_out = srv_conn.exec_command(command)
                while temp_out.find('  ') != -1:
                    temp_out = temp_out.replace('  ', ' ')
                temp_out = temp_out.split('\n')
                temp_result = []
                for line in temp_out:
                    line_out = line.split(' ')
                    temp_result.append(line_out)
                print(temp_result)
                for i in range(0, len(temp_result)):
                    for valu in temp_result[i]:
                        if valu == 'com.ibm.ws.runtime.WsServer':
                            print("couldnt kill the processes")
                            output += "couldn't proceed as some app processes still running\n"
                            break

                if len(was_processes) > 0:
                    output += "\n## deleting temp files\n"
                    print('## deleting temp files')
                    wstemp = srv.bin_path.replace('bin', 'wstemp')
                    temp = srv.bin_path.replace('bin', 'temp')
                    tranlog = srv.bin_path.replace('bin', 'tranlog')
                    workspace = srv.bin_path.replace('bin', 'workspace')
                    command = 'rm -r ' + wstemp + ' ' + temp + ' ' + tranlog + ' ' + workspace
                    output += command
                    output += '\n'
                    output += srv_conn.exec_command(command)

                output += '\n**** Restart Websphere App Server ****\n'
                print('**** Restart Websphere App Server ****')

                for appProcess in was_processes:
                    if appProcess[-1] == 'nodeagent':
                        command = 'cd ' + srv.bin_path + ' ; ' + './startNode.sh'
                        output += '\n'
                        output += srv_conn.exec_command(command)
                        output += '\n'

                for appProcess in was_processes:
                    if appProcess[-1] != 'nodeagent':
                        command = 'cd ' + srv.bin_path + ' ; ' + './startServer.sh ' + appProcess[-1]
                        output += command
                        output += '\n'
                        output += srv_conn.exec_command(command)
                        output += '\n'
            in_queue.put(output)

        except EOFError as e:
            err = 'cannot connect machine : ' + srv.IP
            print(err)
            print(e)
            in_queue.put(err)
        except paramiko.ssh_exception.AuthenticationException:
            err = 'Invalid username or password for machine : ' + srv.IP
            print(err)
            in_queue.put(err)
        except Resource.DoesNotExist:
            print('please link resources to this machine : ' + srv.IP)
            in_queue.put('please link resources to this machine : ' + srv.IP)
        except Exception as e:
            print(e)
            in_queue.put(str(e))
        finally:
            srv_conn.disconnect_ssh()
            return


@csrf_exempt
@permission_required('auth.can_check_archive_requests', login_url="/login", raise_exception=True)
def search_archive_fn(request):
    print(
        '=============================================== New Thread in search_archive_fn =======================================')
    if request.method == 'POST':
        asynch = request.POST['Async']

        if asynch != '':
            messages = []
            if 'search_group' in request.POST:
                search_group_name = request.POST['search_group']
                if request.user.has_perm('auth.retrieve_log_req_res_only') and not request.user.is_staff:
                    search_type = 'archive_msgs'
                elif request.user.has_perm('auth.retrieve_full_log'):
                    search_type = 'archive'
                else:
                    search_type = 'archive'
                log_type = request.POST['log_type']
                search_group = Group.objects.get(groupName=search_group_name)
                matchingIPs = Server.objects.filter(group_name__groupName=search_group_name)
                component_type = Group.objects.get(groupName=search_group).AppComponentTypeCode
                errors = Error.objects.filter(AppComponentType=component_type)
                trx_date = request.POST['trx_date']
                trx_hr = request.POST['trx_time']

                in_queue = queue.Queue()
                files_list = queue.Queue()
                threads_list = list()

                list_of_archive_period_string = search_group.archiveCountPerDay
                list_of_archive_period = list_of_archive_period_string.split(',')

                print(list_of_archive_period)
                list_trx_archive_period = []
                for period in list_of_archive_period:
                    trx_hr_div = int(int(trx_hr) / int(period))
                    trx_archive_period = int(period) * (trx_hr_div + 1)
                    list_trx_archive_period.append(str(trx_archive_period))

                print(list_trx_archive_period)

                for s in matchingIPs:
                    t = threading.Thread(target=show_archived_messages, args=(s, asynch, in_queue, files_list,
                                                                              trx_date, list_trx_archive_period,))
                    t.start()
                    threads_list.append(t)

                for t in threads_list:
                    t.join()

                print("finished handling archived logs threads ===================================")

                while not files_list.empty():
                    filename_statement = files_list.get()
                    filename = filename_statement.split(' ')[-1]
                    hostnameIP = filename_statement.split(' ')[-4]
                    filename = filename.strip()
                    is_file = os.path.isfile( local_win_out_path + "\\" + filename)
                    is_exist = os.path.exists(local_win_out_path + "\\" + filename)
                    if not (is_file & is_exist):
                        print("Cannot open file: " + filename + " on machine: " + hostnameIP)
                        messages.append(["Cannot open file: " + filename + " on machine: " + hostnameIP])
                    else:
                        messages.append(["Found logs in file: " + filename.strip() + " on machine: " + hostnameIP])
                        messages.append(search_file_by_async(filename, asynch, log_type, search_type, errors, component_type))

            t = TemplateResponse(request, 'show_messages.html', {'messages': messages})
            return t.render()
    return HttpResponse("Please enter the required text for search")


@csrf_exempt
@permission_required('auth.can_handle_obie_check', login_url="/login", raise_exception=True)
def obie_handling_fn(request):
    print(
        '=============================================== New Thread in obie_handling_fn =======================================')
    if request.method == 'POST':
        print("In obie_handling_fn")
        output = checkobiehealth()
        print(output)
    #return HttpResponse(output.decode("utf-8"))
    return HttpResponse(output)


@csrf_exempt
@permission_required('auth.can_check_log_requests', login_url="/login", raise_exception=True)
def get_messages_from_logs_fn(request):
    print(
        '=============================================== New Thread in get_messages_from_logs_fn =======================================')
    if request.method == 'POST':
        if 'Async' in request.POST and 'trx_date' in request.POST:
            messages = []
            if 'search_group' in request.POST:
                asynch = request.POST['Async']
                search_group = request.POST['search_group']
                matchingIPs = Server.objects.filter(group_name__groupName=search_group)
                component_type = Group.objects.get(groupName=search_group).AppComponentTypeCode

                if request.user.has_perm('auth.retrieve_log_req_res_only') and not request.user.is_staff:
                    search_type = 'archive_msgs'
                elif request.user.has_perm('auth.retrieve_full_log'):
                    search_type = 'archive'
                else:
                    search_type = 'archive'

                print(search_type)
                errors = Error.objects.filter(AppComponentType=component_type)

                trx_date = request.POST['trx_date']
                trx_hour = ''
                if 'trx_hour' in request.POST:
                    trx_hour = request.POST['trx_hour']
                is_switch = request.POST['is_switch']
                in_queue = queue.Queue()
                files_list = queue.Queue()
                threads_list = list()
                for s in matchingIPs:
                    t = threading.Thread(target=show_messages,
                                         args=(s, asynch, in_queue, files_list, is_switch, trx_date, trx_hour))
                    t.start()
                    threads_list.append(t)

                for t in threads_list:
                    t.join()

                if in_queue.queue[0] == "operation failed duo to exception, please try again":
                    messages.append(["operation failed duo to exception, please try again or contact support"])

                else:
                    while not files_list.empty():
                        filename_statement = files_list.get()
                        filename = filename_statement.split(' ')[-1]
                        hostnameIP = filename_statement.split(' ')[4]
                        if filename_statement.find("Error") == 0:
                            messages.append([filename_statement])
                        elif not (os.path.isfile(local_win_out_path + '\\' + filename) & os.path.exists(local_win_out_path + '\\' + filename)):
                            messages.append(["Cannot open file: " + filename + " in machine " + hostnameIP])
                        else:
                            messages.append([filename_statement])
                            if search_type == 'archive_msgs':
                                result = list()
                                msgs = search_file_by_async(filename, asynch, is_switch, search_type, errors, component_type)

                                if environment_nature in ('staging', 'local'):

                                    date_time_obj = datetime.datetime.strptime(trx_date, '%Y-%m-%d')
                                    year = date_time_obj.strftime("%Y")
                                    month = date_time_obj.strftime("%m")
                                    day = date_time_obj.strftime("%d")

                                    search_date = '[' + day + '/' + month + '/' + year + ' ' + trx_hour
                                    print(search_date)

                                    rendered_msgs = list()

                                    for msg in msgs:
                                        x = msgs.index(msg)
                                        try:
                                            if 'Start Handling new thread with async' in msg or 'Finished Handling new thread with async' in msg:
                                                if search_date in msgs[x + 1]:
                                                    rendered_msgs.append(msg)
                                            elif search_date in msg:
                                                rendered_msgs.append(msg)

                                                for i in range(x + 1, x + 300):
                                                    line = msgs[i]
                                                    if re.match('^\[[0-9]+/[0-9]+/[0-9]+', line):
                                                        break
                                                    else:
                                                        rendered_msgs.append(line)
                                        except IndexError as e:
                                            break

                                if environment_nature in ('staging', 'local'):
                                    logs_error_logic(rendered_msgs, result, component_type)
                                    if len(rendered_msgs) > 0:
                                        messages.append(rendered_msgs)
                                else:
                                    logs_error_logic(msgs, result, component_type)
                                    if len(msgs) > 0:
                                        messages.append(msgs)

                                if len(result) > 0:
                                    messages.append(result)

                            elif search_type == 'archive':
                                messages.append(
                                    search_file_by_async(filename, asynch, is_switch, search_type, errors, component_type))

            t = TemplateResponse(request, 'show_messages.html', {'messages': messages})
            t.render()
            return t.render()

    return HttpResponse("fail")


def show_messages(srv, asynch, out_queue, files_list, is_switch, trx_date, trx_hour):
    try:
        srv_conn = ServerConnection(srv)
        srv_conn.connect_ssh()
        hostname = srv.IP
        user = srv.app_user
        pwd = srv.app_password
        log_path = srv.log_path
        logfile_prefix = srv.logfile_prefix
        tmp = greptool(hostname, log_path, asynch, out_queue, is_switch, trx_date, trx_hour, logfile_prefix, srv_conn)

        if len(tmp) != 0:
            tmp = tmp.decode("UTF-8")
            print("Received files to search for:\n===============================\n" + tmp)
            if tmp.find('Error') != 0:
                deleteTempFiles(hostname, user, pwd, srv_conn)
                tmp = tmp.strip()
                grep_result_list = tmp.split('\n')
                for file in grep_result_list:
                    copyFileToTemp(hostname, user, pwd, log_path, file, srv_conn)
                    if file == "trace.log":
                        files_list.put("Found file in machine: " + hostname + " trace" + hostname + ".log")
                    else:
                        files_list.put("Found file in machine: " + hostname + " " + file)
                move_to_local(srv)
            else:
                files_list.put(tmp)
        else:
            files_list.put("Error: no result return for machine: " + hostname)

    except Exception as e:
        files_list.put("Error: failed to get files for server %s , exception: %s" % (hostname, str(e)))
    finally:
        srv_conn.disconnect_ssh()
    return


def show_archived_messages(server, asynch, in_queue, files_list, trx_date, list_trx_archive_period, ):
    tmp_file_list = greparchivetool(server, asynch, in_queue, trx_date, list_trx_archive_period,)
    if tmp_file_list is not None:
        deleteTempFiles(server.IP, server.app_user, server.app_password)
        for tmp_file in tmp_file_list:
            #tmp_file = tmp_file.decode("utf-8").strip()
            print("Found archived file in server: " + server.IP + " with name: " + tmp_file)
            copyFileToTemp(server.IP, server.app_user, server.app_password, '/waslogs/archivetmp', tmp_file)
            files_list.put("Found archived file in server: " + server.IP + " with name: " + tmp_file)

        move_to_local(server)
    return


@csrf_exempt
@permission_required('auth.can_change_log_level', login_url="/login", raise_exception=True)
def changeLogLevelFn(request):
    print(
        '=============================================== New Thread in changeLogLevelFn =======================================')
    if request.method == 'POST':
        if 'search_group' in request.POST:
            is_open_log = request.POST['is_open_log']
            is_close_log = request.POST['is_close_log']
            search_group = request.POST['search_group']
            matchingIPs = Server.objects.filter(group_name__groupName=search_group)
            in_queue = queue.Queue()
            threads_list = list()
            script_name = ''
            log_level = GroupLoggingLevel.objects.filter(group__groupName=search_group)

            for s in matchingIPs:
                if is_open_log == 'true':
                    script_name = 'control_logging_level.py ' + log_level[0].logs_opened_packages
                elif is_close_log == 'true':
                    script_name = 'control_logging_level.py ' + log_level[0].logs_closed_packages
                print(script_name)
                t = threading.Thread(target=runLoggingLevelThread, args=(s, s.app_user, s.app_password, s.soap_pass, s.soap_port, s.bin_path, in_queue, script_name))
                t.start()
                threads_list.append(t)

            for t in threads_list:
                t.join()

            stdout = ''
            while not in_queue.empty():
                result = in_queue.get()
                stdout += result
                stdout += '\n'
            print(stdout)
            return HttpResponse(stdout)
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.can_get_string_count', login_url="/login", raise_exception=True)
def count_exception_fn(request):
    print(
        '=============================================== New Thread in count_exception_fn =======================================')
    if request.method == 'POST':
        if 'exception_code' in request.POST:
            if 'search_group' in request.POST:
                exception_code = request.POST['exception_code']
                search_group = request.POST['search_group']
                matchingIPs = Server.objects.filter(group_name__groupName=search_group)
                trx_time = request.POST['trx_time']
                trx_date = request.POST['trx_date']
                log_type = request.POST['is_switch']
                in_queue = queue.Queue()
                threads_list = list()
                for s in matchingIPs:
                     t = threading.Thread(target=greptool1, args=(s.IP, s.log_path, exception_code, in_queue, log_type, trx_date, trx_time, s.logfile_prefix,))
                     t.start()
                     threads_list.append(t)

                for t in threads_list:
                     t.join()

                stdout = ''
                while not in_queue.empty():
                    result = in_queue.get()
                    stdout += result
                    stdout += '\n'
                print(stdout)
            return HttpResponse(stdout)
        else:
            HttpResponse("Please choose exception")
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.can_check_using_soap', login_url="/login", raise_exception=True)
def check_service_soap_status_fn(request):
    print('=========================================== New Thread in check_service_soap_status_fn =======================================')
    if request.method == 'POST':
        stdout = ''
        try:
            search_env = request.POST['search_env']
            envObject = Environment.objects.filter(envName=search_env)

            services = Service.objects.all()
            in_queue = queue.Queue()
            threads_list = list()
            for service in services:
                soap_reqs = SoapRequest.objects.filter(req_env__envName=search_env, req_service=service)
                if not soap_reqs:
                    continue

                for req in soap_reqs:

                    serversGroups = req.req_group.all()
                    req = SoapRequest.objects.filter(id=req.id)
                    for group in serversGroups:

                        if envObject[0] in group.environments.all():

                            ipsList = Server.objects.filter(group_name=group)
                            print(ipsList)
                            for ip in ipsList:
                                if ip != '':
                                    srv = Server.objects.get(IP=ip)
                                    t = threading.Thread(target=soapTool, args=(srv.IP, srv.port, group, in_queue, req, service))
                                    t.start()
                                    threads_list.append(t)

            for t in threads_list:
                t.join()

            result = list()
            while not in_queue.empty():
                result.append(in_queue.get())

            stdout = '<html>\n' + '<head>\n' + '<style>\n' + 'table, th, td {\n'
            stdout += 'border: 3px solid black;\n}'
            stdout += '</style>\n</head>\n<body>\n'
            for service in services:
                stdout += '<h4>' + service.name + '</h4>'
                stdout += '=========================='
                soap_reqs = SoapRequest.objects.filter(req_env__envName=search_env, req_service=service)
                for req_entry in soap_reqs:
                    reqCategory = list()
                    for entry in result:
                        if req_entry.name == entry[1]:
                            reqCategory.append([entry[2], entry[3]])
                    stdout += '<h6><b>' + req_entry.name + '</b></h6>'
                    stdout += '<table>\n'
                    stdout += '<tr>\n'
                    for reqCat in reqCategory:
                        tmp = '<th>' + reqCat[0] + '</th>\n'
                        stdout += tmp
                    stdout += '</tr>\n<tr>\n'
                    for reqCat in reqCategory:
                        if 'success' in reqCat[1] or 'Unauthorized' in reqCat[1]:
                            tmp = '<th class=\"thresholdclear\">Success</th>\n'
                        else :
                            tmp = '<th class=\"thresholdalert\">Fail</th>\n'
                        stdout += tmp
                    stdout += '</tr>\n'
                    stdout += '</table>\n'
            stdout += '</body>\n</html>'

        except Exception as e:
            print(str(e))
            traceback.print_exc()
        return HttpResponse(stdout)
    return HttpResponse('error')


@csrf_exempt
@permission_required('auth.can_check_using_soap', login_url="/login", raise_exception=True)
def check_soap_status_fn(request):
    print('=========================================== New Thread in check_soap_status_fn =======================================')
    if request.method == 'POST':
        stdout = ''
        try:
            search_group = request.POST['search_group']
            ip_list_string = request.POST.get('ip_list')
            ip_list = ip_list_string.split(':')
            #print(ip_list)

            #soap_reqs = SoapRequest.objects.filter(req_group__groupName=search_group, req_app__Code=app)
            soap_reqs = SoapRequest.objects.filter(req_group__groupName=search_group)
            if not soap_reqs:
                return HttpResponse("Please add requests for this group to be able to retrieve them")
            in_queue = queue.Queue()
            threads_list = list()

            for ip in ip_list:
                if ip != '':
                    srv = Server.objects.get(IP=ip)
                    t = threading.Thread(target=soapTool, args=(srv.IP, srv.port, search_group, in_queue, soap_reqs, 'REQUEST'))
                    t.start()
                    threads_list.append(t)

            for t in threads_list:
                t.join()

            while not in_queue.empty():
                result = in_queue.get()
                stdout += result
                stdout += '\n'

        except Exception as e:
            print(str(e))
            traceback.print_exc()
        return HttpResponse(stdout)
    return HttpResponse('error')


@csrf_exempt
@permission_required('auth.command_executer', login_url="/login", raise_exception=True)
def execute_command_fn(request):
    print('=========================================== New Thread in execute_command_fn =======================================')
    if request.method == 'POST':
        stdout = ''
        try:
            search_group = request.POST['search_group']
            command = request.POST.get('Async')
            ip_list_string = request.POST.get('ip_list')
            ip_list = ip_list_string.split(':')

            in_queue = queue.Queue()
            threads_list = list()

            for ip in ip_list:
                if ip != '':
                    srv = Server.objects.get(IP=ip)
                    t = threading.Thread(target=cmdTool, args=(srv.IP, command, in_queue))
                    t.start()
                    threads_list.append(t)

            for t in threads_list:
                t.join()

            while not in_queue.empty():
                result = in_queue.get()
                stdout += str(result)
                stdout += '\n'

        except Exception as e:
            print(str(e))
            traceback.print_exc()
        return HttpResponse(stdout)
    return HttpResponse('error')


@csrf_exempt
@permission_required('auth.can_check_env_health', login_url="/login", raise_exception=True)
def check_env_health_fn(request):
    print(
        '=============================================== New Thread in check_env_health_fn =======================================')
    if request.method == 'POST':
        if 'search_env' in request.POST:
            searchEnv = request.POST['search_env']
            minimum_hr_range = request.POST['gt_hour']
            maximum_hr_range = request.POST['lt_hour']
            groups = Group.objects.filter(environments__envName=searchEnv, Status__Code='active')

            in_queue = queue.Queue()
            threads_list = list()
            for g in groups:
                t = threading.Thread(target=env_groups_search, args=(g, in_queue, minimum_hr_range, maximum_hr_range, False))
                t.start()
                threads_list.append(t)

            for t in threads_list:
                t.join()

            stdout = '<html>\n' +'<head>\n' +'<style>\n' +'table, th, td {\n'
            stdout += 'border: 3px solid black;\n}'
            stdout += '</style>\n</head>\n<body>\n'
            stdout += '<h1 style="text-align:center">' + "Current time: " + str(time.ctime()) + '</h1>\n\n'

            while not in_queue.empty():
                result = in_queue.get()
                stdout += result
                stdout += '\n\n\n'

            stdout += '</body>\n</html>'

        return HttpResponse(stdout)
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.check_rep_query_monitor_page', login_url="/login", raise_exception=True)
def check_rep_queries_fn(request):
    print(
        '=============================================== New Thread in check_rep_queries_fn =======================================')
    if request.method == 'POST':
        if 'search_env' in request.POST:
            searchEnv = request.POST['search_env']

            databases = Database.objects.filter(environment__envName=searchEnv)
            in_queue = queue.Queue()
            threads_list = list()
            for db in databases:
                queries = RepresentationalQuery.objects.filter(database=db, environments__envName=searchEnv)
                if len(queries) > 0:
                    t = threading.Thread(target=get_representational_query_result, args=(db, in_queue, searchEnv))
                    t.start()
                    threads_list.append(t)

            for t in threads_list:
                t.join()

            stdout = '<html>\n' +'<head>\n' +'<style>\n' +'table, th, td {\n'
            stdout += 'border: 3px solid black;\n}'
            stdout += '</style>\n</head>\n<body>\n'
            #stdout += '<h1 style="text-align:center">' + "Current time: " + str(time.ctime()) + '</h1>\n\n<div style="display: flex; flex-direction: row;">'
            stdout += '<h1 style="text-align:center">' + "Current time: " + str(
                time.ctime()) + '</h1>\n\n'

            while not in_queue.empty():
                result = in_queue.get()
                stdout += result
                stdout += '\n\n\n'

            stdout += '</body>\n</html>'

        return HttpResponse(stdout)
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.can_monitor_version_downloads', login_url="/login", raise_exception=True)
def download_monitor_fn(request):
    print(
        '=============================================== New Thread in download_monitor_fn =======================================')
    if request.method == 'POST':
        if 'search_group' in request.POST:
            action = request.POST['action']
            search_group = request.POST['search_group']
            matchingIPs = Server.objects.filter(group_name__groupName=search_group)
            trx_date = request.POST['trx_date']
            in_queue = queue.Queue()
            threads_list = list()
            print(matchingIPs)

            for srv in matchingIPs:
                host = Server.objects.get(IP=srv.IP)
                t = threading.Thread(target=countStringTool, args=(host, 'exception_code', in_queue, action, trx_date))
                t.start()
                threads_list.append(t)

            for t in threads_list:
                t.join()

            if action =='get_downloads':
                stdout = []
                server202 = []
                server203 = []
                server204 = []
                hour = []
                while not in_queue.empty():
                    result = in_queue.get()
                    if result[0] == '10.100.43.202':
                        server202 = result
                    if result[0] == '10.100.43.203':
                        server203 = result
                    if result[0] == '10.100.43.204':
                        server204 = result
                    if result[0] == 'hour':
                        hour = result
                stdout.append(hour)
                stdout.append(server202)
                stdout.append(server203)
                stdout.append(server204)

                print(stdout)
                data = []
                for x in range(1,len(stdout[0])):
                    entry = []
                    for y in range(len(stdout)):
                        entry.append(stdout[y][x])
                    data.append(entry)
                print(data)
                return HttpResponse(json.dumps(data))
            elif action == 'get_terminals':
                stdout = []
                while not in_queue.empty():
                    result = in_queue.get()
                    for entry in result:
                        entry[0] = entry[0].replace('], Fawry Pr', '')
                        entry[0] = entry[0].replace('], ', '')
                        entry[0] = entry[0].replace('],', '')
                        entry[0] = entry[0].replace(']', '')

                        terminal_exists = False
                        if len(stdout) > 0:
                            for m in range(len(stdout)):
                                if entry[0] == stdout[m][0]:
                                    stdout[m][1] += entry[1]
                                    terminal_exists == True
                                    break
                        if not terminal_exists:
                            stdout.append(entry)

                filename = "duplicates_downloads.xlsx"
                workbook1 = Workbook()
                sheet1 = workbook1.active
                for row_no in range(len(stdout)):
                    sheet1.cell(row=row_no+1, column=1).value = stdout[row_no][0]
                    sheet1.cell(row=row_no+1, column=2).value = stdout[row_no][1]
                workbook1.save(filename=filename)

                """
                            stdout = '<html>\n' +'<head>\n' +'<style>\n' +'table, th, td {\n'
                            stdout += 'border: 3px solid black;\n}'
                            stdout += '</style>\n</head>\n<body>\n'
                            stdout += '<h1 style="text-align:center">' + "Current time: " + str(time.ctime()) + '</h1>\n\n'
                            while not in_queue.empty():
                                result = in_queue.get()
                                stdout += result
                                stdout += '\n\n\n'

                            stdout += '</body>\n</html>'
                            """
                return HttpResponse("success")
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.check_env_health_graphs', login_url="/login", raise_exception=True)
def check_env_health_graphs_fn(request):
    print(
        '=============================================== New Thread in check_env_health_graphs_fn =======================================')
    if request.method == 'POST':
        search_group = request.POST['search_group']
        trx_date = request.POST['trx_date']
        servers = Server.objects.filter(group_name__groupName=search_group).order_by("IP")

        ips = list()
        for server in servers:
            ips.append(server.IP)

        result = displayExceptionsGraphs(ips, trx_date)

        return HttpResponse(result)
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.check_env_health_graphs', login_url="/login", raise_exception=True)
def check_queues_health_graphs_fn(request):
    print(
        '=============================================== New Thread in check_queues_health_graphs_fn =======================================')
    if request.method == 'POST':
        search_group = request.POST['search_group']
        trx_date = request.POST['trx_date']
        servers = Server.objects.filter(group_name__groupName=search_group).order_by('IP')
        ips = list()
        for server in servers:
            ips.append(server.IP)

        print(ips)
        result = displayQueuesGraphs(ips, trx_date)
        return HttpResponse(result)
    return HttpResponse("fail")


@csrf_exempt
@permission_required('auth.check_queries_health_graphs', login_url="/login", raise_exception=True)
def check_queries_health_graphs_fn(request):
    print(
        '============================================ New Thread in check_queries_health_graphs_fn =======================================')
    if request.method == 'POST':

        database_name = request.POST['database']
        trx_date = request.POST['trx_date']
        severity = request.POST['severity']
        database = Database.objects.get(name=database_name)
        print(database.id, trx_date, severity)
        result = displayQueriesGraphs(database, trx_date, severity)
        return HttpResponse(result)

    return HttpResponse("fail")


@csrf_exempt
def dashboards_fn(request):
    print(
        '=============================================== New Thread in dashboards_fn =======================================')
    if request.method == 'POST':
        action_type = request.POST['action_type']
        dashboard = request.POST['dashboard']
        username = request.POST['username']
        password = request.POST['password']
        user_desc = request.POST['user_desc']
        app_role = request.POST['AppRole']

        print(dashboard)
        try:
            dashboard_object = Dashboard.objects.get(name=dashboard)
            hostname = Server.objects.get(IP='10.100.44.164')

            if dashboard_object.hasAppRole:
                rolename = app_role
            else:
                rolename='test'

            move_to_remote(hostname, 'scripts')

            srv_conn = ServerConnection(hostname)
            srv_conn.connect_ssh()

            command = 'cd ' + hostname.bin_path + ' ; ./wlst.sh ' + remote_aix_deployment_folder + "/obie.py " \
                         + username + ' ' + password + ' ' + user_desc + ' ' + str(dashboard_object.hasAppRole) + ' ' + \
                            rolename

            srv_conn.exec_command(command)

            client = Client('http://' + hostname.IP + ':' + hostname.port + '/analytics-ws/saw.dll/wsdl/v7')
            sessionid = client.service['SAWSessionService'].logon('weblogic', 'P@ssw0rd2015')
            print(sessionid)
            # accountlist = client.factory.create('Account')
            # account = client.factory.create('Account')
            updateCatalogItemACLParams = client.factory.create('UpdateCatalogItemACLParams')
            updateCatalogItemACLParams.updateFlag.value = 1
            updateCatalogItemACLParams.recursive = True

            print(updateCatalogItemACLParams)

            user = client.factory.create('Account')
            user.name = username
            user.accountType = 0

            accounts = client.service['SecurityService'].getAccounts(user, sessionid)

            print(accounts)

            act = client.factory.create('AccessControlToken')
            act.account = accounts[0]
            act.permissionMask = 3

            acl = client.factory.create('ACL')
            acl.accessControlTokens = act

            print(acl)

            permissions = client.service['WebCatalogService'].updateCatalogItemACL(dashboard_object.physicalPath, acl, updateCatalogItemACLParams, sessionid)

            print(permissions)

        except Dashboard.DoesNotExist as e:
            print(str(e))
            traceback.print_exc()
            return HttpResponse("fail")
        except NameError as e:
            print(str(e))
            traceback.print_exc()
            return HttpResponse("fail")
        except Exception as e:
            traceback.print_exc()
            return HttpResponse("fail")

        return HttpResponse('success')
    return HttpResponse("fail")


executors = {
            'default': ThreadPoolExecutor(30),  # max threads: 90
            'processpool': ProcessPoolExecutor(5)  # max processes 20
        }
scheduler = BackgroundScheduler(executors=executors, timezone="Africa/Cairo")


@csrf_exempt
@permission_required('auth.can_control_scheduler_tasks', login_url="/login", raise_exception=True)
def start_sch_fn(request):
    print(
        '=============================================== New Thread in start_sch_fn =======================================')
    if request.method == 'POST':
        action = request.POST['action']


        try:
            if action == 'start-sch':
                scheduler.start()
                return HttpResponse("Scheduler started successfully")
            elif action == 'run-background_env_monitor':
                scheduler.add_job(background_env_monitor, 'interval', minutes=10, id='background_env_monitor')
                return HttpResponse("job background_env_monitor started")
            elif action == 'run-db_queries_checker':
                scheduler.add_job(db_queries_checker, 'interval', minutes=3, id='db_queries_checker')
                return HttpResponse("job db_queries_checker started")
            elif action == 'run-api_requests_checker':
                scheduler.add_job(api_requests_checker, 'interval', minutes=3, id='api_requests_checker')
                return HttpResponse("job api_requests_checker started")
            elif action == 'run-was_queues_checker':
                scheduler.add_job(was_queues_checker, 'interval', minutes=3, id='was_queues_checker')
                return HttpResponse("job was_queues_checker started")
            elif action == 'run-ACH_accounts_job_checker':
                scheduler.add_job(ACH_accounts_job_checker, 'cron', hour=0, minute=5, id='ACH_accounts_job_checker')
                return HttpResponse("job ACH_accounts_job_checker started")
            elif action == 'run-ACH_success_accounts_job_checker':
                scheduler.add_job(ACH_success_accounts_job_checker, 'cron', hour=6, minute=0, id='ACH_success_accounts_job_checker')
                return HttpResponse("job ACH_success_accounts_job_checker started")
            elif action == 'stop-background_env_monitor':
                scheduler.remove_job('background_env_monitor')
                return HttpResponse("job db_queries_checker stopped")
            elif action == 'stop-db_queries_checker':
                scheduler.remove_job('db_queries_checker')
                return HttpResponse("job db_queries_checker stopped")
            elif action == 'stop-api_requests_checker':
                scheduler.remove_job('api_requests_checker')
                return HttpResponse("job api_requests_checker stopped")
            elif action == 'stop-was_queues_checker':
                scheduler.remove_job('was_queues_checker')
                return HttpResponse("job was_queues_checker stopped")
            elif action == 'stop-ACH_accounts_job_checker':
                scheduler.remove_job('ACH_accounts_job_checker')
                return HttpResponse("job ACH_accounts_job_checker stopped")
            elif action == 'stop-ACH_success_accounts_job_checker':
                scheduler.remove_job('ACH_success_accounts_job_checker')
                return HttpResponse("job ACH_success_accounts_job_checker stopped")
            elif action == 'check-jobs-status':
                scheduler.print_jobs()
                print(scheduler.get_jobs())
                return HttpResponse(scheduler.get_jobs())
            elif action == 'shutdown-sch':
                scheduler.shutdown()
                return HttpResponse("Scheduler stopped successfully")
        except schedulers.SchedulerNotRunningError as e:
            print(e)
            return HttpResponse("Scheduler is not running")
        except Exception as e:
            print(e)
            return HttpResponse(e)

        #scheduler.add_interval_job(triggerTask, interval_time, executor="<executor's name>")

    return HttpResponse("fail")


def background_env_monitor():
    print(
        '=============================================== New Thread in background_env_monitor =======================================')
    minimum_hr_range = 0
    maximum_hr_range = 0.25
    groups = Group.objects.filter((~Q(environments__envName='staging')), Status__Code='active', allow_monitoring=True).order_by('groupName')

    threads_list = list()

    for g in groups:

        in_queue = queue.Queue()
        t = threading.Thread(target=env_groups_search, args=(g, in_queue, minimum_hr_range, maximum_hr_range, True))
        t.start()
        threads_list.append(t)

    for t in threads_list:
        t.join()

    return


def db_queries_checker():
    print(
        '=============================================== New Thread in db_queries_checker =======================================')
    queries = Query.objects.all()
    current_time = datetime.datetime.now()
    for q in queries:
        try:
            databases = q.getDatabases()
            databases = databases.split()
            #print(databases)
            threads_list = list()
            for db in databases:
                try:
                    print(db)
                    database = Database.objects.get(name=db)
                    if database.allow_monitoring == True:
                        t = threading.Thread(target=get_query_result, args=(database, q, current_time))
                        t.start()
                        threads_list.append(t)
                    else:
                        continue
                except Exception as e:
                    print(str(e))
                    print("Issue while getting the database info for query")
                    traceback.print_exc()

            for t in threads_list:
                 t.join()

        except Exception as e:
             print(str(e))
             print("Issue while getting the query databases")
             traceback.print_exc()

    return HttpResponse("success")


def api_requests_checker():
    print(
        '============================================== New Thread in api_requests_checker =======================================')
    requests = SoapRequest.objects.filter(is_used_in_job=True)

    try:
        for req in requests:
            threads_list = list()
            groups = req.getGroup()
            groups = groups.split()
            in_queue = queue.Queue()
            for group in groups:
                group_obj = Group.objects.get(groupName=group)
                try:
                    #groupObj = Group.objects.get(groupName=group)
                    #if str(groupObj.Status) == 'inactive':
                    if group in ('PROD-SW-Old'):
                        continue

                    servers = Server.objects.filter(group_name__groupName=group)
                    if len(servers) < 1:
                        continue
                    for server in servers:
                        t = threading.Thread(target=sendSoapReq, args=(server.IP, server.port, group_obj, in_queue, req, True))
                        t.start()
                        threads_list.append(t)
                except Exception as e:
                    print(str(e))
                    print("Issue while getting the request info")
                    traceback.print_exc()

            for t in threads_list:
                t.join()

    except Exception as e:
         print(str(e))
         print("Issue while getting the requests")
         traceback.print_exc()

    return HttpResponse("success")


def was_queues_checker():
    #servers = Server.objects.filter(middleware_tech='Websphere', os__in=['aix','linux'])
    servers = Server.objects.filter(IP='10.100.4.11')
    threads_list = list()
    in_queue = queue.Queue()
    current_time = datetime.datetime.now()
    for server in servers:
        try:
            t = threading.Thread(target=getWebsphereQueueinfo, args=(server, in_queue, current_time, ))
            t.start()
            threads_list.append(t)

        except Exception as e:
            print("issue in server {}: ".format(server.IP) + str(e))
    for t in threads_list:
        t.join()

    result = ''
    while not in_queue.empty():
        result += in_queue.get()

    return HttpResponse(result)


def abnormal_trx_monitor():
    databases = Database.objects.filter(name='PROD-retail-GW-STDBY-DB')
    print(databases)
    for database in databases:
        query = "select * from ebpp_core.payment_transactions where gateway_pmt_creation_date between '03-APR-2023 12:00:00.00000000 AM' and '03-APR-2023 11:59:00.00000000 PM' and payment_status_id in (1,2)"

        if database.sid is not None:
            conn_str = database.IP + ':' + database.port + '/' + database.sid
        else:
            conn_str = database.IP + ':' + database.port + '/' + database.service_name

        try:
            with cx_Oracle.connect(
                    database.user,
                    database.password,
                    conn_str,
                    encoding='UTF-8') as db_conn:
                with db_conn.cursor() as db_conn_cursor:
                    query_rows = []
                    db_conn_cursor.execute(query)
                    while True:
                        db_row = db_conn_cursor.fetchone()
                        if db_row is None:
                            break
                        query_rows.append(db_row)
                    print(query_rows)
                    result = query_rows[0][0]
                    print(result)

        except cx_Oracle.Error as error:
            print("Issue while connecting to " + str(database.name) + " : " + str(error))
        except Exception as e:
            print(str(e))
            traceback.print_exc()


def get_representational_query_result(database, out_queue, searchEnv):

    queries = RepresentationalQuery.objects.filter(database=database, environments__envName=searchEnv).order_by('positioning')

    if database.sid is not None:
        conn_str = database.IP + ':' + database.port + '/' + database.sid
    else:
        conn_str = database.IP + ':' + database.port + '/' + database.service_name

    if database.db_type == 'oracle':
        try:
            with cx_Oracle.connect(
                    database.user,
                    database.password,
                    conn_str,
                    encoding='UTF-8') as db_conn:
                with db_conn.cursor() as db_conn_cursor:
                    positioning = list()
                    view_html = ''
                    for q in queries:
                        counters_string = q.counter.split(';')
                        counters = list()
                        for rec in counters_string:
                            tmp = list()
                            tmp_list = rec.split(',')

                            for item in tmp_list:
                                try:
                                    tmp.append(str(item))
                                except ValueError as e:
                                    tmp.append(item)
                                except e:
                                    print(e)
                            tmp[len(tmp)-1] = int(tmp_list[len(tmp_list)-1])
                            counters.append(tuple(tmp))
                        print('counters', counters)

                        lookups_result = list()
                        query_rows = list()

                        now = datetime.datetime.now()
                        now = now - datetime.timedelta(minutes=q.timer)
                        current_time = now.strftime("%d-%b-%Y %I:%M:%S.00000000 %p")
                        passed = now - datetime.timedelta(minutes=q.timer)
                        old_time = passed.strftime("%d-%b-%Y %I:%M:%S.00000000 %p")
                        query_string = q.Query.format(old_time, current_time)
                        print(query_string)
                        db_conn_cursor.execute(query_string)

                        while True:
                            db_row = db_conn_cursor.fetchone()
                            print(db_row)
                            if db_row is not None:
                                lookups_result.append(db_row[:-1])
                                query_rows.append(db_row)
                            else:
                                break
                        print('query_rows', query_rows)
                        print('lookups_result', lookups_result)

                        new_query_rows = list()
                        if q.is_missing_required == True:
                            for counter_rec in counters:
                                if counter_rec[:-1] not in (lookups_result):
                                    counter_rec = list(counter_rec[:-1])
                                    counter_rec.append(0)
                                    counter_rec.append('missing')
                                    new_query_rows.append(tuple(counter_rec))

                        for row in query_rows:
                            for counter_row in counters:
                                if row[:-1] == counter_row[:-1]:
                                    if str(q.nature) == 'incremental':
                                        if row > counter_row:
                                            row = list(row)
                                            row.append('exceed')
                                            new_query_rows.append(tuple(row))
                                        else:
                                            row = list(row)
                                            row.append('')
                                            new_query_rows.append(tuple(row))

                                    elif str(q.nature) == 'decremental':
                                        if row < counter_row:
                                            row = list(row)
                                            row.append('low')
                                            new_query_rows.append(tuple(row))
                                        else:
                                            row = list(row)
                                            row.append('')
                                            new_query_rows.append(tuple(row))

                        headers = q.headers.split(',')
                        if q.positioning not in positioning:
                            if '<div' not in view_html:
                                view_html = '\n\n<div style="display: flex; flex-direction: row;">'
                            else:
                                view_html += '</div>\n\n<div style="display: flex; flex-direction: row;">'
                            view_html += form_table_entries(q.QueryName, headers, new_query_rows)
                            view_html += '\n\n\n'

                            positioning.append(q.positioning)
                        else:
                            view_html += form_table_entries(q.QueryName, headers, new_query_rows)
                            view_html += '\n\n\n'

                    out_queue.put(view_html + '</div>')

        except cx_Oracle.Error as error:
            print("Issue while connecting to " + str(database.name) + " : " + str(error))
        except Exception as e:
            print(str(e))
            traceback.print_exc()

    elif database.db_type == 'db2':
        now = datetime.datetime.now()
        now = now - datetime.timedelta(minutes=5)
        current_time = now.strftime("%Y-%M-%d %H:%M:%S.00000000")
        passed = now - datetime.timedelta(minutes=5)
        old_time = passed.strftime("%Y-%M-%d %H:%M:%S.00000000")

        print("Current Time =", current_time)
        print("Current Time =", old_time)
        query_string = q.query.format(old_time, current_time)
        print(query_string)
    elif database.db_type == 'sqlserver':
        now = datetime.datetime.now()
        now = now - datetime.timedelta(minutes=5)
        current_time = now.strftime("%d-%b-%Y %I:%M:%S.00000000 %p")
        passed = now - datetime.timedelta(minutes=5)
        old_time = passed.strftime("%d-%b-%Y %I:%M:%S.00000000 %p")

        print("Current Time =", current_time)
        print("Current Time =", old_time)
        query_string = q.query.format(old_time, current_time)
        print(query_string)

    return


def env_groups_search(group, out_queue, min_hr, max_hr, is_background_send_mail):
    print("###inside env_groups_search###")
    matchingIPs = Server.objects.filter(group_name__groupName=group.groupName)
    exceptions = list()

    exceptions_bygroup = Exceptions.objects.filter(Group__groupName=group.groupName, AppComponentType__ComponentTypeCode=None)
    for excep in exceptions_bygroup:
        Exception_Threshold_bygroup = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                             AppComponentType__ComponentTypeCode=None).Threshold
        Warning_Threshold_bygroup = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                           AppComponentType__ComponentTypeCode=None).WarningThreshold
        severity_bygroup = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                           AppComponentType__ComponentTypeCode=None).ExceptionSeverity
        excep_status = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                           AppComponentType__ComponentTypeCode=None).Exception_Status
        count = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                           AppComponentType__ComponentTypeCode=None).count_btn_alerts
        counter = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                           AppComponentType__ComponentTypeCode=None).counter
        interval_btn_alerts = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                           AppComponentType__ComponentTypeCode=None).interval_btn_alerts
        lastAlertTime = Exceptions.objects.get(Code=excep, Group__groupName=group.groupName,
                                                           AppComponentType__ComponentTypeCode=None).lastAlertTime
        exception_entry = (excep, Exception_Threshold_bygroup, Warning_Threshold_bygroup, severity_bygroup, excep_status,
                           count, counter, interval_btn_alerts, lastAlertTime)
        exceptions.append(exception_entry)

    exceptions_bycomponent = Exceptions.objects.filter(AppComponentType__ComponentTypeCode=group.AppComponentTypeCode, Group__groupName=None)
    for excep in exceptions_bycomponent:
        Exception_Threshold_bycomponent = Exceptions.objects.get(Code=excep,
                                                                 AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                                                 Group__groupName=None).Threshold
        Warning_Threshold_bycomponent = Exceptions.objects.get(Code=excep,
                                                               AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                                               Group__groupName=None).WarningThreshold
        severity_bycomponent = Exceptions.objects.get(Code=excep,
                                                               AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                                               Group__groupName=None).ExceptionSeverity
        excep_status = Exceptions.objects.get(Code=excep,
                                                      AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                                      Group__groupName=None).Exception_Status
        count = Exceptions.objects.get(Code=excep,
                                             AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                             Group__groupName=None).count_btn_alerts
        counter = Exceptions.objects.get(Code=excep,
                                              AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                              Group__groupName=None).counter
        interval_btn_alerts = Exceptions.objects.get(Code=excep,
                                              AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                              Group__groupName=None).interval_btn_alerts
        lastAlertTime = Exceptions.objects.get(Code=excep,
                                              AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                              Group__groupName=None).lastAlertTime
        exception_entry = (excep, Exception_Threshold_bycomponent, Warning_Threshold_bycomponent, severity_bycomponent,
                           excep_status, count, counter, interval_btn_alerts, lastAlertTime)

        Exception_not_exist = True
        for excep_entry in exceptions:
            if str(excep) == str(excep_entry[0]):
                Exception_not_exist = False

        if Exception_not_exist:
            exceptions.append(exception_entry)

    group_in_queue = queue.Queue()
    threads_list = list()

    for s in matchingIPs:
        if s.os in ('aix', 'linux'):
            t = threading.Thread(target=ExceptionsCounter, args=(s, group_in_queue, exceptions, min_hr, max_hr))
            t.start()
            threads_list.append(t)

    for t in threads_list:
        t.join()

    guides = ['exception']
    Exception_Thresholds = ['Thresholds']
    Warning_Thresholds = ['Warning']
    ex_names = ['exceptions']
    for ex in exceptions:
        #print(ex)
        guides.append(str(ex[0]))
        Exception_Threshold = ex[1]
        Warning_Threshold = ex[2]
        Exception_Thresholds.append(Exception_Threshold)
        Warning_Thresholds.append(Warning_Threshold)
        ex_names.append(str(ex[0]))

    stdout = []
    stdout.append(guides)

    while not group_in_queue.empty():
        result = group_in_queue.get()
        stdout.append(result)

    table = "<h1>" + group.groupName + "</h1>\n<table>\n"
    current_time = datetime.datetime.now()
    # Create the table's row data
    for y_value in range(0, len(stdout[0])):
        threshold = Exception_Thresholds[y_value]
        warning_threshold = Warning_Thresholds[y_value]
        ex_name = ex_names[y_value]
        if y_value != 0:
            severity = exceptions[y_value-1][3]
            exception_status = exceptions[y_value-1][4]
            count = exceptions[y_value - 1][5]
            #counter = exceptions[y_value - 1][6]
            interval = exceptions[y_value - 1][7]
            lastAlertTime = exceptions[y_value - 1][8]
        #print(ex_name + str(threshold) + str(severity))
        table += "  <tr>\n"
        for x_value in range(0, len(stdout)):
            ipaddress = stdout[x_value][0]
            if re.match('^[0-9]+.[0-9]+.[0-9]+.[0-9]+', ipaddress):
                srv = Server.objects.get(IP=ipaddress)
            try:

                if threshold == 'Thresholds' or threshold is None or warning_threshold == 'Warning' \
                        or warning_threshold is None or not re.match('^[0-9]+', str(stdout[x_value][y_value])):
                    table += "    <td>" + str(stdout[x_value][y_value]) + "</td>\n"

                elif str(severity) == 'critical':
                    if is_background_send_mail and str(exception_status) == 'active':
                        addExceptionMonitorRecord(ex_name, group, srv, stdout[x_value][y_value], current_time)

                        if int(stdout[x_value][y_value]) > int(threshold) / 4:
                            try:
                                send_mail(ex_name, '', group.groupName, str(stdout[x_value][y_value]), threshold,
                                          severity,
                                          'moustafa.mamdouh@fawry.com', 'app', ipaddress)
                            except Exception as e:
                                print("Cannot send mail duo to: " + str(e))
                    elif int(stdout[x_value][y_value]) > int(threshold):
                        table += "    <td class=\"thresholdalert\">" + str(stdout[x_value][y_value]) + "</td>\n"
                    else:
                        table += "    <td>" + str(stdout[x_value][y_value]) + "</td>\n"

                elif str(severity) == 'warning':
                    if is_background_send_mail and str(exception_status) == 'active':
                        addExceptionMonitorRecord(ex_name, group, srv, stdout[x_value][y_value], current_time)

                        if int(stdout[x_value][y_value]) > int(threshold)/4:
                            try:
                                send_mail(ex_name, '', group.groupName, str(stdout[x_value][y_value]), threshold,
                                          severity,
                                          'moustafa.mamdouh@fawry.com', 'app', ipaddress)
                            except TimeoutError as e:
                                print("Cannot send mail duo to timeout")
                    elif int(stdout[x_value][y_value]) > int(warning_threshold) and int(stdout[x_value][y_value]) < int(threshold):
                        table += "    <td class=\"thresholdwarn\">" + str(stdout[x_value][y_value]) + "</td>\n"
                    elif int(stdout[x_value][y_value]) > int(threshold):
                        table += "    <td class=\"thresholdalert\">" + str(stdout[x_value][y_value]) + "</td>\n"
                    else:
                        table += "    <td>" + str(stdout[x_value][y_value]) + "</td>\n"
                elif str(severity) == 'behavioral':
                    if is_background_send_mail and str(exception_status) == 'active':
                        addExceptionMonitorRecord(ex_name, group, srv, stdout[x_value][y_value], current_time)

                        if int(stdout[x_value][y_value]) > int(threshold)/4:
                            exception_bycomponent = Exceptions.objects.filter(Code=ex_name,
                                                                              AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                                                              Group__groupName=None).first()
                            exception_bygroup = Exceptions.objects.filter(Code=ex_name,
                                                                          Group__groupName=group.groupName,
                                                                          AppComponentType__ComponentTypeCode=None).first()
                            if exception_bycomponent == None:
                                exception_counter = \
                                    ExceptionsCounters.objects.filter(Exception_id__id=exception_bygroup.id,
                                                                      server_id__IP=ipaddress).first()
                            else:
                                exception_counter = \
                                    ExceptionsCounters.objects.filter(Exception_id__id=exception_bycomponent.id,
                                                                      server_id__IP=ipaddress).first()
                            if exception_counter == None:
                                counter = 0
                            else:
                                counter = exception_counter.counter

                            if counter == 0:
                                counter += 1
                                try:
                                    send_mail(ex_name, '', group.groupName, str(stdout[x_value][y_value]),
                                              threshold,
                                              severity,
                                              'moustafa.mamdouh@fawry.com', 'app', ipaddress)
                                except TimeoutError as e:
                                    print("Cannot send mail duo to timeout")
                            else:
                                if count <= counter:
                                    counter = 0
                                else:
                                    counter += 1

                            if exception_bycomponent == None:
                                exception_counter = \
                                    ExceptionsCounters.objects.filter(Exception_id__id=exception_bygroup.id,
                                                                      server_id__IP=ipaddress).first()
                                if exception_counter == None:
                                    counterObject = ExceptionsCounters(
                                        Exception_id=exception_bygroup,
                                        server_id=srv,
                                        counter=counter,
                                        count_btn_alerts=count,
                                        interval_btn_alerts=interval,
                                        lastAlertTime=datetime.datetime.now()
                                    )
                                    counterObject.save()
                                else:
                                    exception_counter.counter= counter
                                    exception_counter.save()
                            elif exception_bygroup == None:
                                exception_counter = \
                                    ExceptionsCounters.objects.filter(Exception_id__id=exception_bycomponent.id,
                                                                      server_id__IP=ipaddress).first()
                                print(exception_counter)
                                if exception_counter == None:
                                    counterObject = ExceptionsCounters(
                                        Exception_id=exception_bycomponent,
                                        server_id=srv,
                                        counter=counter,
                                        count_btn_alerts=count,
                                        interval_btn_alerts=interval,
                                        lastAlertTime=datetime.datetime.now()
                                    )
                                    counterObject.save()
                                else:
                                    exception_counter.counter = counter
                                    exception_counter.save()
                    elif int(threshold) > int(stdout[x_value][y_value]) > int(warning_threshold):
                        table += "    <td class=\"thresholdwarn\">" + str(stdout[x_value][y_value]) + "</td>\n"
                    elif int(stdout[x_value][y_value]) > int(threshold):
                        table += "    <td class=\"thresholdalert\">" + str(stdout[x_value][y_value]) + "</td>\n"
                    else:
                        table += "    <td>" + str(stdout[x_value][y_value]) + "</td>\n"
                elif str(severity) == 'configurational':
                    if is_background_send_mail and str(exception_status) == 'active':
                        addExceptionMonitorRecord(ex_name, group, srv, stdout[x_value][y_value], current_time)

                        if int(stdout[x_value][y_value]) > int(threshold)/4:
                            currenttime = datetime.datetime.now()
                            diff = currenttime - lastAlertTime
                            if diff.total_seconds() > interval * 60:
                                exception_bycomponent = Exceptions.objects.filter(Code=ex_name,
                                                                                  AppComponentType__ComponentTypeCode=group.AppComponentTypeCode,
                                                                                  Group__groupName=None).first()
                                exception_bygroup = Exceptions.objects.filter(Code=ex_name,
                                                                              Group__groupName=group.groupName,
                                                                              AppComponentType__ComponentTypeCode=None).first()
                                if exception_bycomponent == None:
                                    exception_bygroup.lastAlertTime = datetime.datetime.now()
                                    exception_bygroup.save()
                                elif exception_bygroup == None:
                                    exception_bycomponent.lastAlertTime = datetime.datetime.now()
                                    exception_bycomponent.save()

                                try:
                                    send_mail(ex_name, '', group.groupName, str(stdout[x_value][y_value]),
                                              threshold,
                                              severity,
                                              'moustafa.mamdouh@fawry.com', 'app', ipaddress)
                                except TimeoutError as e:
                                    print("Cannot send mail duo to timeout")
                    elif int(threshold) > int(stdout[x_value][y_value]) > int(warning_threshold):
                        table += "    <td class=\"thresholdwarn\">" + str(stdout[x_value][y_value]) + "</td>\n"
                    elif int(stdout[x_value][y_value]) > int(threshold):
                        table += "    <td class=\"thresholdalert\">" + str(stdout[x_value][y_value]) + "</td>\n"
                    else:
                        table += "    <td>" + str(stdout[x_value][y_value]) + "</td>\n"
            except ValueError as e:
                table += "    <td>" + str(stdout[x_value][y_value]) + "</td>\n"
            except Exception as e:
                print("Exception " + str(e))
        table += "  </tr>\n"

    table += "</table>"
    out_queue.put(table)
    return table


def ExceptionsCounter(server, group_out_queue, exceptions, min_hr, max_hr):

    username = server.app_user
    password = server.app_password

    mini_min = 60 * int(min_hr)
    max_min = 60 * float(max_hr)

    try:
        cmd1 = 'cd ' + server.log_path + ' ; ' + 'find * -mmin +' + str(mini_min) + ' -mmin -' + str(int(max_min))
        port = 22


        client = paramiko.SSHClient()
        client.load_system_host_keys()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(server.IP, port, username, password)
        channel = client.get_transport().open_session()
        channel.exec_command(cmd1)
        while not channel.exit_status_ready():
            time.sleep(1)
        stdout = channel.makefile("rb")
        temp = stdout.readlines()

        if temp:
            grep_result_list = []
            for rec in temp:
                tmp = rec.decode("utf-8").strip()
                grep_result_list.append(tmp)

            counters = [server.IP]

            for ex in exceptions:
                temp_str = server.IP + "\n======================\n" + cmd1 + "\n"
                e_count = 0
                grouped_command = ''

                for f in grep_result_list:
                    cmd2 = 'cd ' + server.log_path + ' ; ' + 'grep -c "' + str(
                        ex[0]) + '" ' + f + ' ; '
                    grouped_command += cmd2

                temp_str += grouped_command

                channel = client.get_transport().open_session()
                channel.exec_command(grouped_command)
                while not channel.exit_status_ready():
                    time.sleep(1)
                stdout = channel.makefile("rb")
                tmp = stdout.readlines()

                for rec in tmp:
                    temp = rec.decode("utf-8").strip()
                    try:
                        e_count += int(temp)
                    except ValueError:
                        continue

                counters.append(e_count)

            group_out_queue.put(counters)
        return
    except EOFError as e:
        print('Exception in ' + server.IP + ': ' + str(e))
        group_out_queue.put('Exception in ' + server.IP + ': ' + cmd2)
        return
    except TimeoutError as e:
        print('Exception in ' + server.IP + ': ' + str(e))
        group_out_queue.put("Connection Timeout: Unable to connect")
        return
    except paramiko.ssh_exception.AuthenticationException:
        print("Unable to connect to: " + server.IP + " Please check the connection credentials.\n")
    except Exception as e:
        print('Exception in ' + server.IP + ': ' + str(e))
        return
    finally:
        client.close()


def runLoggingLevelThread(hostname, username, password, soap_pass, soap_port, was_bin_path, out_queue, script_name):
    move_to_remote(hostname, 'scripts')
    execute_script(hostname.IP, username, password, soap_pass, soap_port, was_bin_path, out_queue, script_name)


def ServerCreateForm(request):
    form = ServerForm(request.POST or None)
    if form.is_valid():
        form.save()

    context = {
        'form': form
    }
    return render(request, 'Server_create.html', context)


def groupCreateForm(request):
    form = GroupForm(request.POST or None)
    if form.is_valid():
        form.save()

    context = {
        'form': form
    }
    return render(request, 'Group_Create.html', context)


def ServerCreatePureForm(request):
    my_form = ServerPureForm()
    if request.method == 'POST':
        my_form = ServerPureForm(request.POST)
        if my_form.is_valid():
            print(my_form.cleaned_data)
            Server.objects.create(**my_form.cleaned_data)
        else:
            print(my_form.errors)

    context = {
        'form': my_form
    }
    return render(request, 'Server_Pure_create.html', context)


def greptool(hostname, path, asynch, out_queue, log_type, trx_date, trx_hour, logfile_prefix, srv_conn=None):
    try:
        username = username_property
        password = password_property

        if log_type == 'switch':
            date_time_obj = datetime.datetime.strptime(trx_date, '%Y-%m-%d')
            year = date_time_obj.strftime("%y")
            month = date_time_obj.strftime("%m")
            if int(month) < 10:
                month = month.replace('0', '')
            day = date_time_obj.strftime("%d")
            if int(day) < 10:
                day = day.replace('0', '')

            if environment_nature == 'staging':
                command = "cd " + path + " ; grep -l '" + asynch + "' switch_" + year + "*"
            else:
                command = "cd " + path + " ; grep -l '" + asynch + "' switch_" + year + "." + month + "." + day + "_" + trx_hour + "*"


        elif log_type == 'mcc-server':
            command = "cd " + path + " ; grep -l '" + asynch + "' *"

        elif log_type == 'mcc-sof':
            command = "cd " + path + " ; grep -l '" + asynch + "' *"

        elif log_type == 'mcc-gateway':
            command = "cd " + path + " ; grep -l '" + asynch + "' *"

        elif log_type == 'sw-iib' and environment_nature != 'staging':
            username = 'mqm'
            password = 'P@ssw0rd'
            date_time_obj = datetime.datetime.strptime(trx_date, '%Y-%m-%d')
            year = date_time_obj.strftime("%y")
            month = date_time_obj.strftime("%m")
            if int(month) < 10:
                month = month.replace('0', '')
            day = date_time_obj.strftime("%d")
            if int(day) < 10:
                day = day.replace('0', '')


            if logfile_prefix == None:
                command = "cd " + path + " ; grep -l '" + asynch + "' switch_" \
                      + year + "." + month + "." + day + "_" + trx_hour + "*"
            elif hostname in ('10.100.4.78', '10.100.4.88'):
                command = "cd " + path + " ; grep -l '" + asynch + "' " + logfile_prefix + "_switch_" \
                          + year + "." + month + ".*"
            else:
                command = "cd " + path + " ; grep -l '" + asynch + "' " + logfile_prefix + "_switch_" \
                      + year + "." + month + "." + day + "_" + trx_hour + "*"
        else:
            command = "cd " + path + " ; grep -l '" + asynch + "' *"

        port = 22

        if srv_conn is not None:
            temp = srv_conn.exec_command(command)
            temp = temp.encode("UTF-8")
        else:
            srv = Server.objects.get(IP=hostname)
            srv_conn = ServerConnection(srv)
            srv_conn.connect_ssh()
            temp = srv_conn.exec_command(command)
            temp = temp.encode("utf-8")
            #client = paramiko.SSHClient()
            #client.load_system_host_keys()
            #client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            #client.connect(hostname, port, username, password)
            #client.exec_command(command)
            #stdin, stdout, stderr = sr.exec_command(command)
            #temp = stdout.read()

        out_queue.put(hostname + "\n======================\n" + temp.decode("utf-8"))
        return temp
    except EOFError as e:
        print(e)
        error_msg = "Error: " + hostname + "\n======================\n" + "EOF"
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")
    except TimeoutError as e:
        print(e)
        error_msg = "Error: " + hostname + "\n======================\n" + "Connection Timeout: Unable to connect"
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")
    except Exception as e:
        error_msg = "Error: " + "Exception in server: " + hostname + ' ' + str(e)
        print(error_msg)
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")
    finally:
        if srv_conn is None:
            srv_conn.disconnect_ssh()
            #client.close()


def cmdTool(hostname, command, out_queue, srv_conn=None):
    try:
        username = username_property
        password = password_property
        port = 22

        if srv_conn is not None:
            temp = srv_conn.exec_command(command)
            temp = temp.encode("UTF-8")
        else:
            srv = Server.objects.get(IP=hostname)
            srv_conn = ServerConnection(srv)
            srv_conn.connect_ssh()
            temp = srv_conn.exec_command(command)
            temp = temp.encode("utf-8")
        out_queue.put(hostname + "\n=====================\n" + temp.decode("utf-8"))
        return temp
    except EOFError as e:
        print(e)
        error_msg = "Error: " + hostname + "\n======================\n" + "EOF"
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")
    except TimeoutError as e:
        print(e)
        error_msg = "Error: " + hostname + "\n======================\n" + "Connection Timeout: Unable to connect"
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")
    except Exception as e:
        error_msg = "Error: " + "Exception in server: " + hostname + ' ' + str(e)
        print(error_msg)
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")
    finally:
        if srv_conn is None:
            srv_conn.disconnect_ssh()


def countStringTool(srv, asynch, out_queue, action, trx_date, srv_conn=None):
    try:

        srv_conn = ServerConnection(srv)
        srv_conn.connect_ssh()
        if action == 'get_downloads':
            date_time_obj = datetime.datetime.strptime(trx_date, '%Y-%m-%d')
            year = date_time_obj.strftime("%y")
            month = date_time_obj.strftime("%m")
            if int(month) < 10:
                month = month.replace('0', '')
            day = date_time_obj.strftime("%d")
            if int(day) < 10:
                day = day.replace('0', '')

            srv_hr_info = [srv.IP]
            hours = ['hour']
            counts = []
            for hr in range(0,24):
                command = "cd " + srv.log_path + " ; grep -n '\\[" + month + '/' + day + '/' + year + " " + str(hr) + "' trace* | gawk '{print $2}' FS=\":\""
                temp = srv_conn.exec_command(command, False).splitlines()
                #srv_hr_info.append((hr, temp[0], temp[-1]))

                if len(temp) > 0:
                    #command = "cd " + srv.log_path + " ; sed -n '" + temp[0] + "," + temp[-1] + "p' trace* | grep \"========= Start Downlod Service\""
                    command = "cd " + srv.log_path + " ; sed -n '" + temp[0] + "," + temp[
                        -1] + "p' trace* | grep \"getDownloadURL downloadedPhysicalDirPath = /home/wasadmin/apps/FRMA/POS/3.8.1\""

                    temp = srv_conn.exec_command(command).splitlines()
                    counts.append((str(hr), len(temp)))

            for count in counts:
                hours.append(count[0])
                srv_hr_info.append(count[1])

            out_queue.put(srv_hr_info)
            out_queue.put(hours)
        elif action == 'get_terminals':
            terminals = []
            command = "cd " + srv.log_path + ' ; grep "WapDownloadServlet doPost  1- Start Downlod Service for" trace* | gawk \'{print $4}\' FS="["'
            temp = srv_conn.exec_command(command, True).splitlines()
            print(len(temp))

            for line in temp:
                terminal = line[0:22]
                terminal_exists = False
                if len(terminals) > 0:
                    for m in range(len(terminals)):
                        if terminal == terminals[m][0]:
                            terminals[m][1] += 1
                            terminal_exists = True
                            break

                if terminal_exists == False:
                    terminals.append([terminal, 1])
            print(terminals)
            out_queue.put(terminals)
        return
    except EOFError as e:
        print(e)
        error_msg = "Error: " + srv.IP + "\n======================\n" + "EOF"
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")
    except TimeoutError as e:
        print(e)
        error_msg = "Error: " + srv.IP + "\n======================\n" + "Connection Timeout: Unable to connect"
        out_queue.put(error_msg)
        return error_msg.encode("utf-8")

    finally:
        srv_conn.disconnect_ssh()


def checkobiehealth():
    username = "oracle"
    password = "oracle"

    command = "cd /obi/obiee117/wlserver_10.3/common/bin ; ./wlst.sh /obi/New.py"
    #command = "cd /obi/obiee117/wlserver_10.3/common/bin ; pwd;"
    print("10.100.44.164" + ': ' + command)
    port = 22

    try:
        client = paramiko.SSHClient()
        client.load_system_host_keys()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect('10.100.44.164', port, username, password)
        channel = client.get_transport().open_session()
        channel.exec_command(command)
        while not channel.exit_status_ready():
            time.sleep(1)

        stdout = channel.makefile("rb")
        temp = stdout.readlines()
        print(stdout)
        return temp
    except EOFError as e:
        print(e)
        return b'EOF'
    except TimeoutError as e:
        print(e)
        return b'Connection Timeout: Unable to connect'
    finally:
        client.close()


def greparchivetool(srv, asynch, out_queue, trx_date , list_trx_archive_period,):
    #username = srv.app_user
    #password = srv.app_password

    date_time_obj = datetime.datetime.strptime(trx_date, '%Y-%m-%d')
    year = date_time_obj.strftime("%y")
    month = date_time_obj.strftime("%m")
    day = date_time_obj.strftime("%d")

    commands = ''
    #commands = commands + 'cd /waslogs/log ; rm *.log *.tar *.gz *.log.*'
    commands = commands + 'cd /waslogs/temp ; rm *.* ;' + ' cd /waslogs/archivetmp ; rm *.* '

    for hour in list_trx_archive_period:
        day_tmp = day
        month_tmp = month
        year_tmp = year
        if hour == '24':
            hour = '00'
            day_no = int(day) + 1
            if day_no >= 31:
                month_no = int(month) + 1
                if month in ['04', '06', '09', '11']:
                    day_tmp = '01'
                    if month_no < 10:
                        month_tmp = ("0" + str(month_no))
                    else:
                        month_tmp = str(month_no)
                elif month in ['01', '03', '05', '07', '08', '10', '12']:
                    if day_no == 32:
                        day_tmp = '01'
                        if month_no == 13:
                            month_tmp = '01'
                            year_no = int(year_no) + 1
                            year_tmp = str(year_no)
                        elif month_no < 10:
                            month_tmp = ("0" + str(month_no))
                        else:
                            month_tmp = str(month_no)
                    else:
                        day_tmp = '31'

            elif day_no < 10:
                day_tmp = ("0" + str(day_no))
            elif month == '02' and day_no == 29:
                day_tmp = '01'
                month_tmp = '03'
            else:
                day_tmp = str(day_no)
        elif hour in ['0','1','2','3','4','5','6','7','8','9']:
            hour = '0' + hour

        if srv.archive_mode == 'default':
            commands = commands + '; cd ' + srv.archive_path
            commands = commands + '; cp archived_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '-' + hour + ':00.tar.gz /waslogs/archivetmp'
            commands = commands + '; cd /waslogs/archivetmp ; ' + 'gunzip archived_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '-' + hour + ':00.tar.gz'
            commands = commands + '; tar -xvf archived_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '-' + hour + ':00.tar'

        elif srv.archive_mode == 'tracelog_YYYY-MM-DD-HH.MM.SS':
            commands = commands + '; cd ' + srv.archive_path
            commands = commands + '; cp tracelog_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar.gz /waslogs/archivetmp'
            commands = commands + '; cd /waslogs/archivetmp ; ' + 'gunzip tracelog_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar.gz'
            commands = commands + '; tar -xvf tracelog_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar'

        elif srv.archive_mode == 'archived_YYYY-MM-DD-HH.MM.SS':
            commands = commands + '; cd ' + srv.archive_path
            commands = commands + ' ; cp archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar.gz /waslogs/archivetmp'
            commands = commands + ' ; cd /waslogs/archivetmp ; ' + 'gunzip archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar.gz'
            commands = commands + ' ; tar -xvf archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar'

        elif srv.archive_mode == 'tracelog_DD-MM-YYYY_HH.MMpm/am':
            commands = commands + '; cd ' + srv.archive_path
            if int(hour) > 11:
                commands = commands + '; cp tracelog_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '_' + hour + '.00pm.tar.gz /waslogs/archivetmp'
                commands = commands + '; cd /waslogs/archivetmp ; ' + 'gunzip tracelog_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '_' + hour + '.00pm.tar.gz'
                commands = commands + '; tar -xvf tracelog_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '_' + hour + '.00pm.tar'
            else:
                commands = commands + '; cp tracelog_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '_' + hour + '.00am.tar.gz /waslogs/archivetmp'
                commands = commands + '; cd /waslogs/archivetmp ; ' + 'gunzip tracelog_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '_' + hour + '.00am.tar.gz'
                commands = commands + '; tar -xvf tracelog_' + day_tmp + '-' + month_tmp + '-20' + year_tmp + '_' + hour + '.00am.tar'

        elif srv.archive_mode == 'archived_YYYY-MM-DD-HH.MM.SS':
            commands = commands + '; cd ' + srv.archive_path
            commands = commands + '; cp archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar.gz /waslogs/archivetmp'
            commands = commands + '; cd /waslogs/archivetmp ; ' + 'gunzip archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar.gz'
            commands = commands + '; tar -xvf archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.00.tar'

        elif srv.archive_mode == 'archived_YYYY-MM-DD-HH.MM':
            commands = commands + '; cd ' + srv.archive_path
            commands = commands + '; cp archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.tar.gz /waslogs/archivetmp'
            commands = commands + '; cd /waslogs/archivetmp ; ' + 'gunzip archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.tar.gz'
            commands = commands + '; tar -xvf archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.tar'

        elif srv.archive_mode == 'archived_YYYY-MM-DD-HH':
            commands = commands + '; cd ' + srv.archive_path
            commands = commands + '; cp archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.tar.gz /waslogs/archivetmp'
            commands = commands + '; cd /waslogs/archivetmp ; ' + 'gunzip archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.tar.gz'
            commands = commands + '; tar -xvf archived_20' + year_tmp + '-' + month_tmp + '-' + day_tmp + '-' + hour + '.tar'

        else:
            out_queue.put(srv.IP + "\n======================\n" + "Please choose archiving mode\n")
    #grep_cmd = 'cd /waslogs/log ; grep -l "' + asynch + '" *.log*'
    grep_cmd = 'cd /waslogs/archivetmp ; grep -l "' + asynch + '" *.log*'
    ##commands = commands
    port = 22
    print(commands)

    try:
        srv_conn = ServerConnection(srv)
        srv_conn.connect_ssh()

        if srv.os in ('aix', 'linux'):
            command = 'if [ -d "/waslogs/archivetmp" ]; then echo "archive folder exists"; else mkdir -p /waslogs/archivetmp && chmod 777 /waslogs/archivetmp ;  fi'
            srv_conn.exec_command(command)
            command = 'if [ -d "/waslogs/temp" ]; then echo "temp folder exists"; else mkdir -p /waslogs/temp && chmod 777 /waslogs/temp ;  fi'
            srv_conn.exec_command(command)

        elif host.os == 'windows':
            command = '"if exist "' + remote_win_deployment_folder + '\\archivetmp" (echo ""scripts exist"") else (mkdir ' + \
                      remote_win_deployment_folder + '\\archivetmp)"'
            srv_conn.exec_command(command)

        temp = srv_conn.exec_command(commands)
        print(temp)

        temp = srv_conn.exec_command(grep_cmd)
        print(temp)
        if temp:
            return temp.splitlines()
        else:
            out_queue.put("<b>" + srv.IP + "\n=====================</b>\n" + "\tno result found\n")
            return
        """
        client = paramiko.SSHClient()
        client.load_system_host_keys()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(srv.IP, port, username, password)
        channel = client.get_transport().open_session()
        
        if srv.os in ('aix', 'linux'):
            command = 'if [ -d "/waslogs/archivetmp" ]; then echo "archive folder exists"; else mkdir -p /waslogs/archivetmp && chmod 777 /waslogs/archivetmp ;  fi'
            channel.exec_command(command)
            command = 'if [ -d "/waslogs/temp" ]; then echo "temp folder exists"; else mkdir -p /waslogs/temp && chmod 777 /waslogs/temp ;  fi'
            channel.exec_command(command)

        elif host.os == 'windows':
            command = '"if exist "' + remote_win_deployment_folder + '\\archivetmp" (echo ""scripts exist"") else (mkdir ' + \
                      remote_win_deployment_folder + '\\archivetmp)"'
            channel.exec_command(command)
        
        channel.exec_command(commands)
        while not channel.exit_status_ready():
            time.sleep(1)
        stdout = channel.makefile("rb")
        temp = stdout.readlines()
        print(temp)

        channel = client.get_transport().open_session()
        channel.exec_command(grep_cmd)
        while not channel.exit_status_ready():
            time.sleep(1)
        stdout = channel.makefile("rb")
        temp = stdout.readlines()
        print(temp)
        if temp:
            return temp
        else:
            out_queue.put("<b>" + srv.IP + "\n=====================</b>\n" + "\tno result found\n")
            return
            
        """
    except EOFError as e:
        print(e)
        out_queue.put(srv.IP + "\n======================\n" + "EOF\n")
        return
    except TimeoutError as e:
        print(e)
        out_queue.put(srv.IP + "\n======================\n" + "Connection Timeout: Unable to connect\n")
        return
    except paramiko.ssh_exception.AuthenticationException as e:
        print(e)
        out_queue.put(srv.IP + "\n======================\n" + "Connection Failed: Authentication issue\n")
        return srv.IP + "\n======================\n" + "Connection Failed: Authentication issue\n"
    finally:
        srv_conn.disconnect_ssh()



def greptool1(hostname, path, asynch, out_queue, log_type, trx_date, trx_time, logfile_prefix):
    username = username_property
    password = password_property
    print(log_type)
    date_time_obj = datetime.datetime.strptime(trx_date, '%Y-%m-%d')
    year = date_time_obj.strftime("%y")
    month = date_time_obj.strftime("%m")
    day = date_time_obj.strftime("%d")

    if log_type == 'switch':
        if int(month) < 10:
            month = month.replace('0', '')
        if int(day) < 10:
            day = day.replace('0', '')
        command = "cd " + path + " ; grep -c '" + asynch + "' switch_" + year + "." + month + "." + day + "_" + trx_time + "*"
    elif log_type == 'mcc-server':
        command = "cd " + path + " ; grep -c '" + asynch + "' server*"
    elif log_type == 'mcc-sof':
        command = "cd " + path + " ; grep -c '" + asynch + "' sof.log*"
    elif log_type == 'mcc-gateway':
        command = "cd " + path + " ; grep -c '" + asynch + "' gateway.log*"
    elif log_type == 'sw-iib':
        username = 'mqm'
        password = 'P@ssw0rd'
        if int(month) < 10:
            month = month.replace('0', '')
        if int(day) < 10:
            day = day.replace('0', '')
        command = "cd " + path + " ; grep -c '" + asynch + "' " + logfile_prefix + "_switch_" + year + "." + month + "." + day + "_" + trx_time + "*"
    else:
        command = "cd " + path + " ; grep -c '" + asynch + "' trace_" + year + "." + month + "." + day + "_" + trx_time + "* trace.log"

    print(hostname + ': ' + command)
    port = 22

    try:
        client = paramiko.SSHClient()
        client.load_system_host_keys()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(hostname, port, username, password)
        client.exec_command(command)
        stdin, stdout, stderr = client.exec_command(command)
        temp = stdout.read()
        print(temp)
        out_queue.put(hostname + "\n======================\n" + temp.decode("utf-8"))
        return temp
    except EOFError as e:
        print(e)
        out_queue.put(hostname + "\n======================\n" + "EOF")
        return b''
    except TimeoutError as e:
        print(e)
        out_queue.put(hostname + "\n======================\n" + "Connection Timeout: Unable to connect")
        return b''
    finally:
        client.close()


def execute_script(hostname, user, passwrd, soap_pass, soap_port, was_bin_path, out_queue, script_name):
    username = user
    password = passwrd

    command = "cd " + was_bin_path + " ; ./wsadmin.sh -lang jython -conntype SOAP -port " + soap_port + " -username "\
              + user + " -password " + soap_pass + " -f " + remote_aix_deployment_folder + "/scripts/" + script_name
    print(command)
    port = 22

    try:
        client = paramiko.SSHClient()
        client.load_system_host_keys()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(hostname, port, username, password)
        client.exec_command(command)
        stdin, stdout, stderr = client.exec_command(command)
        temp = stdout.read()
        out_queue.put(hostname + "\n======================\n" + temp.decode("utf-8"))
        return temp
    except EOFError:
        return b''
    finally:
        client.close()


def grepfiletool(hostname, path, name):

    username = username_property
    password = password_property
    cnopts = pysftp.CnOpts()
    cnopts.hostkeys = None
    print(path)
    try:
        with pysftp.Connection(hostname, username=username, password=password ,cnopts=cnopts) as sftp:
            for entry in sftp.listdir_attr('/waslogs'):
                if entry.st_mode == 33184:#3320433184
                    sftp.cd('/waslogs')
                    #print(entry.filename)
                    if entry.filename == 'native_stderr.log':
                        sftp.get("native_stderr.log", 'D:\\test\\err.log', preserve_mtime=False)

                #remotepath = '/waslogs' + "/" + entry.filename
                #localpath = os.path.join('D:\\TempLogs', entry.filename)


            #with sftp.cd('/waslogs'):  # temporarily chdir to allcode
            #    #sftp.put('/pycode/filename')  # upload file to allcode/pycode on remote
            #    sftp.get(name)  # get a remote file

        return 'success'
    finally:
        sftp.close()


def move_to_local(srv):
    hostIP = srv.IP

    if hostIP in ['11.96.0.127', '11.96.0.40']:
        username = 'mqm'
        password = 'P@ssw0rd'
    else:
        username = username_property
        password = password_property

    print("In move_to_local from server: " + hostIP)
    filepath = batch_files_path + r'\transfer.bat'
    try:
        process = subprocess.Popen([filepath, hostIP, username, password], shell=True, stdout=subprocess.PIPE)
        (outs, errs) = process.communicate(input=None, timeout=None)
    except Exception as e:
        print(str(e))
        traceback.print_exc()
    finally:
        process.terminate()

    print(outs)
    return outs


def move_to_remote(host, type):
    print("In move_to_remote")

    if type == 'scripts':
        try:
            srv_conn = ServerConnection(host)
            srv_conn.connect_ssh()

            if host.os in ('aix', 'linux'):
                command = 'if [ -d "' + remote_aix_deployment_folder + '/scripts" ]; then echo "scripts exists"; else mkdir -p ' + \
                          remote_aix_deployment_folder + '/scripts && chmod 777 ' + remote_aix_deployment_folder + '/scripts;  fi'
                srv_conn.exec_command(command)
                filepath = batch_files_path + r'\transfer_remote.bat'
                return srv_conn.move_remotely_to_win(filepath)

            elif host.os == 'windows':
                command = '"if exist "' + remote_win_deployment_folder + '\scripts" (echo ""scripts exist"") else (mkdir ' + \
                          remote_win_deployment_folder + '\scripts)"'
                srv_conn.exec_command(command)
                return srv_conn.move_remotely_to_win('scripts')

        except paramiko.ssh_exception.AuthenticationException:
            err = 'Invalid username or password for machine : ' + host.IP
            return err.encode("utf-8")
        except Exception as e:
            print(str(e))
            return b''
        finally:
            srv_conn.disconnect_ssh()

    elif type == 'ear' or type == 'sw_ear':
        try:
            srv_conn = ServerConnection(host)
            srv_conn.connect_ssh()

            if host.os in ('aix', 'linux'):
                command = 'if [ -d "' + remote_aix_deployment_folder + '/ear" ]; then echo "old ear exists" && rm -r ' + \
                          remote_aix_deployment_folder + '/ear/* ; else mkdir -p ' + \
                          remote_aix_deployment_folder + '/ear && chmod 777 ' + remote_aix_deployment_folder + '/ear;  fi'
                srv_conn.exec_command(command)
                if type == 'ear':
                    filepath = batch_files_path + r'\transfer_ear_remote.bat'
                elif type == 'sw_ear':
                    filepath = batch_files_path + r'\transfer_sw_ear_remote.bat'
                return srv_conn.move_remotely_to_win(filepath)
            elif host.os == 'windows':
                command = '"if exist "' + remote_win_deployment_folder + r'\ear" (echo ""old ear exists"" && del /q ' + \
                          remote_win_deployment_folder + r'\ear) else (mkdir ' + \
                          remote_win_deployment_folder + r'\ear)"'
                srv_conn.exec_command(command)
                return srv_conn.move_remotely_to_win('ear')

        except paramiko.ssh_exception.AuthenticationException:
            err = 'Invalid username or password for machine : ' + host.IP
            return err.encode("utf-8")
        except Exception as e:
            print(str(e))
            return b''
        finally:
            srv_conn.disconnect_ssh()

    elif type == 'resources':
        try:
            srv_conn = ServerConnection(host)
            srv_conn.connect_ssh()

            if host.os in ('aix', 'linux'):
                command = 'if [ -d "' + remote_aix_deployment_folder + '/resources" ]; then echo "resources exists" && rm -r ' + \
                          remote_aix_deployment_folder + '/resources/* ; else mkdir -p ' + \
                          remote_aix_deployment_folder + '/resources && chmod 777 ' + remote_aix_deployment_folder + '/resources;  fi'
                srv_conn.exec_command(command)

                command = 'if [ -d "' + remote_aix_deployment_folder + '/backup" ]; then echo "resources exists" ; else mkdir -p ' + \
                          remote_aix_deployment_folder + '/backup && chmod 777 ' + remote_aix_deployment_folder + '/backup;  fi'
                srv_conn.exec_command(command)

                filepath = batch_files_path + r'\transfer_resources_remote.bat'
                out = srv_conn.move_remotely_to_win(filepath)

                client = paramiko.SSHClient()
                client.load_system_host_keys()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(host.IP, 22, username_property, password_property)
                command = 'chmod -R 777 /waslogs/deployment/resources/*'
                stdin, stdout, stderr = client.exec_command(command)
                temp = stdout.read()
                if temp != None:
                    print('%%Changing mode of resources files%%')
                    #print(temp.decode("utf-8"))
                client.close()
                return out
            elif host.os == 'windows':
                command = '"if exist "' + remote_win_deployment_folder + r'\resources" (echo ""resources exists"" && del /q ' + \
                          remote_win_deployment_folder + r'\resources && FOR /D %p IN ("' + remote_win_deployment_folder +\
                          r'\resources\*.*") DO rmdir "%p") else (mkdir ' + \
                          remote_win_deployment_folder + r'\resources)"'
                srv_conn.exec_command(command)
                return srv_conn.move_remotely_to_win('resources')

        except paramiko.ssh_exception.AuthenticationException:
            err = 'Invalid username or password for machine : ' + host.IP
            return err.encode("utf-8")
        except Exception as e:
            print(str(e))
            return b''
        finally:
            srv_conn.disconnect_ssh()

    elif type == 'lib':
        try:
            srv_conn = ServerConnection(host)
            srv_conn.connect_ssh()

            if host.os in ('aix', 'linux'):
                command = 'if [ -d "' + remote_aix_deployment_folder + '/lib" ]; then echo "lib exists" && rm -r ' + \
                          remote_aix_deployment_folder + '/lib/* ; else mkdir -p ' + \
                          remote_aix_deployment_folder + '/lib && chmod 777 ' + remote_aix_deployment_folder + '/lib;  fi'
                srv_conn.exec_command(command)

                filepath = batch_files_path + r'\transfer_lib_remote.bat'
                out = srv_conn.move_remotely_to_win(filepath)

                client = paramiko.SSHClient()
                client.load_system_host_keys()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(host.IP, 22, username_property, password_property)
                command = 'chmod -R 777 /waslogs/deployment/lib/*'
                stdin, stdout, stderr = client.exec_command(command)
                temp = stdout.read()
                if temp != None:
                    print('%%Changing mode of resources files%%')
                    #print(temp.decode("utf-8"))
                client.close()
                return out
            elif host.os == 'windows':
                command = '"if exist "' + remote_win_deployment_folder + r'\lib" (echo ""lib exists"" && del /q ' + \
                          remote_win_deployment_folder + r'\lib && FOR /D %p IN ("' + remote_win_deployment_folder +\
                          r'\lib\*.*") DO rmdir "%p") else (mkdir ' + \
                          remote_win_deployment_folder + r'\lib)"'
                srv_conn.exec_command(command)
                return srv_conn.move_remotely_to_win('lib')

        except paramiko.ssh_exception.AuthenticationException:
            err = 'Invalid username or password for machine : ' + host.IP
            return err.encode("utf-8")
        except Exception as e:
            print(str(e))
            return b''
        finally:
            srv_conn.disconnect_ssh()

    elif type == 'queues':
        try:
            srv_conn = ServerConnection(host)
            srv_conn.connect_ssh()
            if host.os in ('aix', 'linux'):
                command = 'if [ -d "' + remote_aix_deployment_folder + '/queues" ]; then echo "queues exists" && rm -r ' + \
                          remote_aix_deployment_folder + '/queues/* ; else mkdir -p ' + \
                          remote_aix_deployment_folder + '/queues && chmod 777 ' + remote_aix_deployment_folder + '/queues;  fi'
                srv_conn.exec_command(command)
                filepath = batch_files_path + r'\transfer_queues_remote.bat'
                return srv_conn.move_remotely_to_win(filepath)
            elif host.os == 'windows':
                command = '"if exist "' + remote_win_deployment_folder + r'\queues" (echo ""queues exists"" && del /q ' + \
                          remote_win_deployment_folder + r'\queues && FOR /D %p IN ("' + remote_win_deployment_folder +\
                          r'\queues\*.*") DO rmdir "%p") else (mkdir ' + \
                          remote_win_deployment_folder + r'\queues)"'
                srv_conn.exec_command(command)
                return srv_conn.move_remotely_to_win('queues')

        except paramiko.ssh_exception.AuthenticationException:
            err = 'Invalid username or password for machine : ' + host.IP
            return err.encode("utf-8")
        except Exception as e:
            print(str(e))
            return b''
        finally:
            srv_conn.disconnect_ssh()

    else:
        print("incorrect type defined in move to remote fn")
        return b'nothing moved. Please check the type in move to remote'


def copyFileToTemp(hostname, user, pwd, path, filename, srv_conn=None):
    print("Thread in copyFileToTemp in server: " + hostname)

    if( filename.strip() == 'trace.log'):
        command = "cp " + path + "/" + filename.strip() + " /waslogs/temp" +"; cd /waslogs/temp ; mv trace.log trace" + hostname + ".log"
    else:
        command = "cp -p " + path + "/" + filename.strip() + " /waslogs/temp ; chmod 777 /waslogs/temp/" + filename.strip()
    print(command)
    port = 22

    try:
        if srv_conn is not None:
            temp = srv_conn.exec_command(command)
            temp = temp.encode("UTF-8")
        else:
            client = paramiko.SSHClient()
            client.load_system_host_keys()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(hostname, port, user, pwd)
            stdin, stdout, stderr = client.exec_command(command)
            temp = stdout.read()
        return temp
    except EOFError as e:
        print("EOF Exception in server: " + hostname)
        print(e)
        return "EOF Exception in server: " + hostname
    except TimeoutError as e:
        print("Connection Timeout: Unable to connect to server: " + hostname)
        print(e)
        return "Connection Timeout: Unable to connect to server: " + hostname
    except paramiko.ssh_exception.SSHException as e:
        print("Unable to copy file " + filename + " from machine: " + hostname)
        print(e)
        return "Unable to copy file " + filename + " from machine: " + hostname
    finally:
        if srv_conn is None:
            client.close()


def deleteTempFiles(hostname, username, password, srv_conn=None):
    print("Thread in deleteTempFiles in server: " + hostname)

    command = "cd /waslogs/temp ; rm *.log *.log.* ; "

    print(command)
    port = 22

    try:
        if srv_conn is not None:
            temp = srv_conn.exec_command(command)
            temp.encode("UTF-8")
        else:
            client = paramiko.SSHClient()
            client.load_system_host_keys()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(hostname, port, username, password)
            stdin, stdout, stderr = client.exec_command(command)
            temp = stdout.read()
        return temp
    except EOFError as e:
        print(e)
        return "EOF Exception in server: " + hostname
    except TimeoutError as e:
        print(e)
        return "Connection Timeout: Unable to connect to server: " +hostname
    finally:
        if srv_conn is None:
            client.close()

"""
def search_file_by_async(filename, Async, is_switch_log):
    messages = []
    print("In search_fe_by_async
    with open("D:\\test\\" + filename, encoding='Latin-1') as f:
        print("D:\\test\\" + filename)
        lines = f.readlines()
        if is_switch_log == 'switch':
            for i in range(0, len(lines)):
                line = lines[i]
                if Async in line:
                    threadline = line[26:34]
                    threadstart= i
                    if 'The Biller Outgoing Request is' in line:
                        messages.append(lines[i+1])
                    break
            for i in range(threadstart, len(lines)):
                line = lines[i]
                if threadline in line:
                    if 'The select configuration from ''BILL_TYPE_CONFIGURATION' in line:
                        messages.append(line)
                    elif 'The Biller Outgoing Request is' in line:
                        messages.append(lines[i + 1])
                    elif 'evaluate :  The Biller Incomming Response is' in line:
                        messages.append(lines[i + 1])
                        break
                    elif 'logException' in line:
                        messages.append(lines[i])
                        for x in range(i+1, i+15):
                            print(x)
                            line= lines[x]
                            if (threadline in line) or (line[0] == '[') :
                                break
                            else:
                                messages.append(lines[x])
                        break
                    else:
                        continue
        else:
            for line in lines:
                if Async in line:
                    if 'Gateway recived message' in line:
                        messages.append(line)
                        print(line)
                    elif 'Message sent to SourceOfFund' in line:
                        messages.append(line)
                        print(line)
                    elif 'Message received from SourceOfFund' in line:
                        messages.append(line)
                        print(line)
                    elif 'Async Message sent to SourceOfFund' in line:
                        messages.append(line)
                        print(line)
                    elif 'Message sent to Switch' in line:
                        messages.append(line)
                        print(line)
                    elif 'Message received from Switch' in line:
                        messages.append(line)
                        print(line)
                    elif 'Gateway sent message' in line:
                        messages.append(line)
                        print(line)
                    elif 'Adapter received message :' in line:
                        messages.append(line)
                        print(line)
                    else:
                        continue
    return messages
"""

def search_file_by_async(filename, Async, is_switch_log, search_type, error_list, component_type):
    messages = []
    print("In search_file_by_async")
    print(local_win_out_path + '\\' + filename)
    try:
     with open(local_win_out_path + '\\' + filename, encoding='Latin-1') as f:
        lines = f.readlines()
        if str(component_type) == 'Switch':
            sw_threads = []
            for i in range(0, len(lines)):
                line = lines[i]
                if Async in line:
                    if 'Start Handling the message with AsyncUID' in line or 'Start Handling a new MIN Receive' in line or 'Sending SMS from : Fawry to mobile number' in line or 'Start Handling a new SMS Integration Flow request with Msg_ID' in line or 'Start Handling a new SMS Integration Flow request' in line:
                        threadline = line[26:34]
                        thread_exists = False
                        if len(sw_threads) > 0:
                            for m in range(len(sw_threads)):
                                if threadline == sw_threads[m][0]:
                                    thread_exists = True
                                    break
                        if not thread_exists:
                            sw_threads.append([threadline, i, Async])

                    elif 'A new Payment notification request for notifying the System' in line:
                        print('just ignore')

                    elif 'Start Handling a new PmtAdvRq from AdviceQueue' in line:
                        threadline = line[26:34]
                        thread_exists = False
                        if len(sw_threads) > 0:
                            for m in range(len(sw_threads)):
                                if threadline == sw_threads[m][0]:
                                    thread_exists = True
                                    break
                        if not thread_exists:
                            sw_threads.append([threadline, i, Async])

                    elif 'The Biller Outgoing Request is' in line:
                        threadline = line[26:34]
                        thread_exists = False
                        if len(sw_threads) > 0:
                            for m in range(len(sw_threads)):
                                if threadline == sw_threads[m][0]:
                                    thread_exists = True
                                    break
                        for x in reversed(range(i)):
                            sub_line = lines[x]
                            if threadline in sub_line:
                                if 'Start Handling the message with AsyncUID' in sub_line:
                                    asyncRqUidIndex = line.index('AsyncUID')
                                    asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]
                                    if not thread_exists:
                                        sw_threads.append([threadline, x, asyncRqUid])
                                        break
                    else:
                        if re.match('^\[[0-9]{2}/[0-9]{2}/[0-9]+', line):
                            threadline = line[26:34]
                        else:
                            prev_line = lines[i-1]
                            threadline = prev_line[26:34]

                        thread_exists = False
                        if len(sw_threads) > 0:
                            for m in range(len(sw_threads)):
                                if threadline == sw_threads[m][0]:
                                    thread_exists = True
                                    break
                        if not thread_exists:
                            for x in reversed(range(i)):
                                sub_line = lines[x]
                                if threadline in sub_line:
                                    if 'Start Handling the message with AsyncUID' in sub_line:
                                        asyncRqUidIndex = sub_line.index('AsyncUID')
                                        asyncRqUid = sub_line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]
                                        sw_threads.append([threadline, x, asyncRqUid])
                                        break

            if len(sw_threads) > 0:
                print(sw_threads)
                print('finished getting threads\n================================')
                for l in range(len(sw_threads)):
                    messages.append(f"#####################Start Handling new thread with async: {search_type} {sw_threads[l][2]}####################")
                    for i in range(sw_threads[l][1], len(lines)):
                        line = lines[i]
                        if sw_threads[l][0] in line:
                            if search_type == 'archive_msgs':
                                """
                                if 'Start Handling the message with AsyncUID' in line:
                                    asyncIndex = line.index('=')
                                    if is_switch_log == 'sw-iib':
                                        newAsyncReqId = line[asyncIndex + 3:asyncIndex + 39]
                                    else:
                                        newAsyncReqId = line[asyncIndex + 2:asyncIndex + 38]
                                    print(newAsyncReqId)
                                    if newAsyncReqId != sw_threads[l][2]:
                                        break
                                    else:
                                        messages.append(line)
                                """
                                if 'Payment Transaction Object has been removed' in line or \
                                        'The Process has been ended for the message with' in line or \
                                        'ExternalSwitchUtil I navigateToExternalSwitchFlow' in line or \
                                        'ErrorReportProcessor I evaluate : Error Report process ended' in line:
                                    messages.append(
                                        f"######################Finished Handling new thread with async: {sw_threads[l][2]}####################")
                                    break
                                else:
                                    for error in error_list:
                                        #err_string = getattr(error, 'ErrorString')
                                        #if err_string in line:
                                        if error.ErrorString in line:
                                            messages.append(line)
                                            for x in range(i + 1, i + 300):
                                                line = lines[x]
                                                if (sw_threads[l][0] in line) or re.match('^\[[0-9]{2}/[0-9]{2}/[0-9]+', line):
                                                    break
                                                else:
                                                    messages.append(lines[x])
                                            break

                            elif search_type == 'archive':
                                if 'Start Handling the message with AsyncUID' in line or 'Sending SMS from : Fawry to mobile number' in line or 'Start Handling a new SMS Integration Flow request with Msg_ID' in line:
                                    asyncIndex = line.index('=')
                                    if is_switch_log == 'sw-iib':
                                        newAsyncReqId = line[asyncIndex + 3:asyncIndex + 39]
                                    else:
                                        newAsyncReqId = line[asyncIndex + 2:asyncIndex + 38]
                                    print(newAsyncReqId)
                                    if newAsyncReqId != sw_threads[l][2]:
                                        break
                                    else:
                                        messages.append(line)
                                elif 'Payment Transaction Object has been removed' in line or \
                                        'The Process has been ended for the message with' in line or \
                                        'ExternalSwitchUtil I navigateToExternalSwitchFlow' in line:
                                    messages.append(line)
                                    messages.append(
                                        f"######################Finished Handling new thread with async: {sw_threads[l][2]}####################")
                                    break
                                else:
                                    messages.append(line)
                                    for x in range(i + 1, i + 300):
                                        line = lines[x]
                                        if (sw_threads[l][0] in line) or re.match('^\[[0-9]{2}/[0-9]{2}/[0-9]{4}', line):
                                            break
                                        else:
                                            messages.append(lines[x])

        elif str(component_type) == 'mcc-eme':
            mcc_threads = []
            for i in range(0, len(lines)):
                line = lines[i]
                if Async in line:
                    if 'ClientSessionHandler:' in lines[i-1]:
                        mccThreadStart = lines[i-1].index('[WebContainer')
                        mccThread = lines[i-1][mccThreadStart:mccThreadStart+18]
                        req_uid_index = line.index('requestId=')
                        req_uid = line[req_uid_index+10:req_uid_index+45]
                        thread_exists = False
                        if len(mcc_threads) > 0:
                            for m in range(len(mcc_threads)):
                                if (mccThread == mcc_threads[m][0]) or (req_uid == mcc_threads[m][2]):
                                    thread_exists = True
                                    break

                        if not thread_exists:
                            mcc_threads.append([mccThread, i-1, req_uid])
                    else:
                        if lines[i-1].find('[WebContainer') != -1:
                            mccThreadStart = lines[i-1].index('[WebContainer')
                            mccThread = lines[i-1][mccThreadStart:mccThreadStart + 18]
                        if line.find('requestId=') == -1:
                            continue
                        else:
                            req_uid_index = line.index('requestId=')
                            req_uid = line[req_uid_index + 10:req_uid_index + 45]
                            thread_exists = False
                            if len(mcc_threads) > 0:
                                for m in range(len(mcc_threads)):
                                    if (mccThread == mcc_threads[m][0]) or (req_uid == mcc_threads[m][2]):
                                        thread_exists = True
                                        break

                            if not thread_exists:
                                for x in reversed(range(i)):
                                    sub_line = lines[x]
                                    if mccThread in sub_line:
                                        if 'ClientSessionHandler:' in sub_line:
                                            mcc_threads.append([mccThread, x, req_uid])
                                            break
            if len(mcc_threads)> 0 :
                print('finished getting threads\n================================')
                for l in range(len(mcc_threads)):
                    for x in range(mcc_threads[l][1], len(lines)):
                        line = lines[x]
                        if mcc_threads[l][0] in line:
                            if 'INFO HTTPFrontEndServlet:' in line:
                                messages.append(line)
                                messages.append(lines[x+1])
                                messages.append('============================================================================================' +
                                    '============================================================================================')
                                messages.append(
                                    '============================================================================================' +
                                    '============================================================================================')
                                break

                            messages.append(line)
                            while '[WebContainer' not in lines[x+1]:
                                messages.append(lines[x+1])
                                x += 1
        elif str(component_type) == 'Promoapp':
            promo_threads = []
            for i in range(0, len(lines)):
                line = lines[i]
                if Async in line:
                    threadline = line[38:55]
                    print(threadline)
                    thread_exists = False
                    if len(promo_threads) > 0:
                        for m in range(len(promo_threads)):
                            if threadline == promo_threads[m][0] and i < (promo_threads[m][1] + 100):
                                thread_exists = True
                                break
                    if not thread_exists:
                        for x in reversed(range(i)):
                            sub_line = lines[x]
                            if threadline in sub_line:
                                if 'Request started' in sub_line:
                                    promo_threads.append([threadline, x])
                                    break
            if len(promo_threads) > 0:
                print('finished getting threads\n================================')

                for l in range(len(promo_threads)):
                    messages.append(
                        f"######################Start Handling new thread with thread uid: {promo_threads[l][0]}####################")
                    for x in range(promo_threads[l][1], len(lines)):
                        line = lines[x]

                        if promo_threads[l][0] in line:
                            if 'Request ended' in line:
                                messages.append(line)
                                messages.append(
                                    f"######################Finished Handling new thread with async: {promo_threads[l][0]}####################")
                                print('thread ended and about to enter new thread. Processing Stopped')
                                break

                            if search_type == 'archive':
                                messages.append(line)
                                for i in range(x + 1, x + 300):
                                    line = lines[i]
                                    if (promo_threads[l][0] in line) or re.match('^[0-9]{4}-[a-zA-Z]{3}-[0-9]{2}', line):
                                        break
                                    else:
                                        messages.append(lines[i])
                            elif search_type == 'archive_msgs':
                                for error in error_list:
                                    if error.ErrorString in line:
                                        messages.append(line)
                                        for i in range(x + 1, x + 300):
                                            line = lines[i]
                                            if (promo_threads[l][0] in line) or re.match('^[0-9]{4}-[a-zA-Z]{3}-[0-9]{2}', line):
                                                break
                                            else:
                                                messages.append(lines[i])
                                        break

        else:
            gw_threads = []
            sof_threads = []
            adaptor_threads = []
            for i in range(0, len(lines)):
                line = lines[i]
                if Async in line:
                    if 'Gateway recieved message :' in line or 'Gateway Received Message' in line or 'Adapter received message' in line or 'Adapter received request' in line:
                        startIndex = line.find(']')
                        gwthreadline = line[startIndex+2 :startIndex+10]
                        try:
                            RqUidIndex = line.index('<RqUID>')
                            RqUid = line[RqUidIndex + 8:RqUidIndex + 44]
                            asyncRqUid = ''
                            if line.find('<AsyncRqUID>') != -1:
                                asyncRqUidIndex = line.index('<AsyncRqUID>')
                                asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]
                        except ValueError as e:
                            RqUidIndex = line.index('RqUID":"')
                            RqUid = line[RqUidIndex + 8:RqUidIndex + 44]
                            asyncRqUid = ''
                            if line.find('AsyncRqUID":"') != -1:
                                asyncRqUidIndex = line.index('AsyncRqUID":"')
                                asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]

                        thread_exists = False
                        if len(gw_threads) > 0:
                            for m in range(len(gw_threads)):
                                if (asyncRqUid == gw_threads[m][3]) or (RqUid == gw_threads[m][2]):
                                    thread_exists = True
                                    break

                        gwthreadstart = i
                        if not thread_exists:
                            gw_threads.append([gwthreadline, gwthreadstart, RqUid, asyncRqUid])

                    elif 'Message sent to SourceOfFund:' in line:
                        startIndex = line.find(']')
                        gwthreadline = line[startIndex + 2:startIndex + 10]
                        trxRqUidIndex = line.index('<TrnRqUID>')
                        trxRqUid = line[trxRqUidIndex+10:trxRqUidIndex+46]

                        asyncRqUidIndex = line.index('<AsyncRqUID>')
                        asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]

                        for z in range(0, len(lines)):
                            line = lines[z]
                            if trxRqUid in line:
                                gwthreadstart = z
                                break

                        thread_exists = False
                        if len(gw_threads) > 0:
                            for m in range(len(gw_threads)):
                                if (asyncRqUid == gw_threads[m][3]) or (trxRqUid == gw_threads[m][2]):
                                    thread_exists = True
                                    break

                        if not thread_exists:
                            gw_threads.append([gwthreadline, gwthreadstart, trxRqUid, asyncRqUid])

                    elif 'Gateway sent message :' in line or 'Gateway Sent Message :' in line:
                        try:
                            startIndex = line.find(']')
                            gwthreadline = line[startIndex + 2:startIndex + 10]
                            trxRqUidIndex = line.index('<RqUID>')
                            trxRqUid = line[trxRqUidIndex+7:trxRqUidIndex+43]
                            asyncRqUid = ''
                            if line.find('<AsyncRqUID>') != -1:
                                asyncRqUidIndex = line.index('<AsyncRqUID>')
                                asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]

                        except ValueError as e:
                            startIndex = line.find(']')
                            gwthreadline = line[startIndex + 2:startIndex + 10]
                            trxRqUidIndex = line.index('RqUID":"')
                            trxRqUid = line[trxRqUidIndex + 7:trxRqUidIndex + 43]
                            asyncRqUid = ''
                            if line.find('AsyncRqUID":"') != -1:
                                asyncRqUidIndex = line.index('AsyncRqUID":"')
                                asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]

                        for z in range(0, len(lines)):
                            line = lines[z]
                            if trxRqUid in line:
                                gwthreadstart = z
                                break

                        thread_exists = False
                        if len(gw_threads) > 0:
                            for m in range(len(gw_threads)):
                                if (asyncRqUid == gw_threads[m][3]) or (trxRqUid == gw_threads[m][2]):
                                    thread_exists = True
                                    break

                        if not thread_exists:
                            gw_threads.append([gwthreadline, gwthreadstart, trxRqUid, asyncRqUid])

                    elif 'SOF:PosMessageConsumer onMessage the XML incoming request' in line:
                        startIndex = line.find(']')
                        softhreadline = line[startIndex + 2:startIndex + 10]
                        softhreadstart = i
                        RqUidIndex = line.index('<RqUID>')
                        RqUid = line[RqUidIndex + 7:RqUidIndex + 43]

                        asyncRqUidIndex = line.index('<AsyncRqUID>')
                        asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]

                        msgCodeStartIndex = line.index('<MsgCode>')
                        msgCodeEndIndex = line.index('</MsgCode>')
                        message_code = line[msgCodeStartIndex + 9:msgCodeEndIndex]

                        thread_exists = False
                        if len(sof_threads) > 0:
                            for m in range(len(sof_threads)):
                                if ((asyncRqUid == sof_threads[m][3]) or (RqUid == sof_threads[m][2])) and message_code == sof_threads[m][4]:
                                    thread_exists = True
                                    break

                        if not thread_exists:
                            sof_threads.append([softhreadline, softhreadstart, RqUid, asyncRqUid, message_code])


                    elif 'SOFBusinessFacadePortBindingImpl process <' in line:
                        startIndex = line.find(']')
                        softhreadline = line[startIndex + 2:startIndex + 10]
                        softhreadstart = i
                        RqUidIndex = line.index('<RqUID>')
                        RqUid = line[RqUidIndex + 7:RqUidIndex + 43]

                        asyncRqUidIndex = line.index('<AsyncRqUID>')
                        asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]

                        msgCodeStartIndex = line.index('<MsgCode>')
                        msgCodeEndIndex = line.index('</MsgCode>')
                        message_code = line[msgCodeStartIndex + 9:msgCodeEndIndex]

                        thread_exists = False
                        if len(sof_threads) > 0:
                            for m in range(len(sof_threads)):
                                if ((asyncRqUid == sof_threads[m][3]) or (RqUid == sof_threads[m][2])) and message_code == sof_threads[m][4]:
                                    thread_exists = True
                                    break

                        if not thread_exists:
                            sof_threads.append([softhreadline, softhreadstart, RqUid, asyncRqUid, message_code])

                    elif 'The incoming request: MsgCode =PmtNotifyRq , AsyncUID' in line or 'process The incoming request: MsgCode =PmtAddRq' in line:
                        startIndex = line.find(']')
                        softhreadline = line[startIndex + 2:startIndex + 10]
                        softhreadstart = i
                        RqUidIndex = line.index('ReqUID')
                        RqUid = line[RqUidIndex + 8:RqUidIndex + 44]

                        asyncRqUidIndex = line.index('AsyncUID')
                        asyncRqUid = line[asyncRqUidIndex + 11:asyncRqUidIndex + 47]

                        msgCodeStartIndex = line.index('<MsgCode>')
                        msgCodeEndIndex = line.index('</MsgCode>')
                        message_code = line[msgCodeStartIndex + 9:msgCodeEndIndex]

                        thread_exists = False
                        if len(sof_threads) > 0:
                            for m in range(len(sof_threads)):
                                if ((asyncRqUid == sof_threads[m][3]) or (RqUid == sof_threads[m][2])) and message_code == sof_threads[m][4]:
                                    thread_exists = True
                                    break

                        if not thread_exists:
                            sof_threads.append([softhreadline, softhreadstart, RqUid, asyncRqUid, message_code])

                    elif 'AdapterExecuter executeFramework Request <' in line:
                        print("matched this line *************")
                        startIndex = line.find(']')
                        softhreadline = line[startIndex + 2:startIndex + 10]
                        softhreadstart = i
                        RqUidIndex = line.index('<RqUID>')
                        RqUid = line[RqUidIndex + 7:RqUidIndex + 43]

                        asyncRqUidIndex = line.index('<AsyncRqUID>')
                        asyncRqUid = line[asyncRqUidIndex + 12:asyncRqUidIndex + 48]

                        msgCodeStartIndex = line.index('<MsgCode>')
                        msgCodeEndIndex = line.index('</MsgCode>')
                        message_code = line[msgCodeStartIndex + 9:msgCodeEndIndex]

                        thread_exists = False
                        if len(adaptor_threads) > 0:
                            for m in range(len(adaptor_threads)):
                                if ((asyncRqUid == adaptor_threads[m][3])  or (RqUid == adaptor_threads[m][2])) and message_code == adaptor_threads[m][4]:
                                    thread_exists = True
                                    break

                        if not thread_exists:
                            adaptor_threads.append([softhreadline, softhreadstart, RqUid, asyncRqUid, message_code])


            if len(gw_threads) > 0 or len(sof_threads) > 0 or len(adaptor_threads) > 0:
                print('finished getting threads\n================================')

                for l in range(len(gw_threads)):
                    messages.append("############################################")
                    messages.append("########### Gateway thread ####################")
                    for x in range(gw_threads[l][1], len(lines)):
                        line = lines[x]

                        if gw_threads[l][0] in line:
                            if 'LoggingSOAPHandler - Gateway sent message :' in line or 'Adapter sent message :' in line or 'Adapter sent response' in line:
                                messages.append(line)
                                print('thread ended and about to enter new thread. Processing Stopped')
                                break

                            if search_type == 'archive':
                                messages.append(line)
                                for i in range(x + 1, x + 300):
                                    line = lines[i]
                                    if (gw_threads[l][0] in line) or re.match('^\[[0-9]+/[0-9]+/[0-9]+', line):
                                        break
                                    else:
                                        messages.append(lines[i])
                            elif search_type == 'archive_msgs':
                                for error in error_list:
                                    if error.ErrorString in line:
                                        messages.append(line)
                                        for i in range(x + 1, x + 300):
                                            line = lines[i]
                                            if (gw_threads[l][0] in line) or re.match('^\[[0-9]+/[0-9]+/[0-9]+', line):
                                                break
                                            else:
                                                messages.append(lines[i])
                                        break

                for l in range(len(sof_threads)):
                    messages.append("##########################################")
                    messages.append("################ SOF thread ###############")

                    for y in range(sof_threads[l][1], len(lines)):
                        line = lines[y]

                        if sof_threads[l][0] in line:
                            if ('onMessage the XML marshaled response: <?xml' in line) or \
                                    ('AdapterExecuter executeFramework Response <?xml' in line) or \
                                    ('PosSofBusines I SOF:PosSofBusinessImpl updateDebit  debitId:' in line) or \
                                    ('><Response>' in line):
                                messages.append(line)
                                break

                            if ('SOF:FawryReqResUtil generateErrorResponse' in line) or 'SOF:MerchantEJB process The sent response: MsgCode =' in line:
                                messages.append(line)
                                break

                            if search_type == 'archive':
                                messages.append(line)
                                for x in range(y + 1, y + 300):
                                    line = lines[x]
                                    if (sof_threads[l][0] in line) or re.match('^\[[0-9]+/[0-9]+/[0-9]+', line):
                                        break
                                    else:
                                        messages.append(lines[x])
                            elif search_type == 'archive_msgs':
                                for error in error_list:
                                    if error.ErrorString in line:
                                        messages.append(line)
                                        for x in range(y + 1, y + 300):
                                            line = lines[x]
                                            if (sof_threads[l][0] in line) or re.match('^\[[0-9]+/[0-9]+/[0-9]+', line):
                                                break
                                            else:
                                                messages.append(lines[x])
                                        break
                for l in range(len(adaptor_threads)):
                    messages.append("##########################################")
                    messages.append("############## Adaptor Log #################")
                    for y in range(adaptor_threads[l][1], len(lines)):
                        line = lines[y]

                        if adaptor_threads[l][0] in line:
                            if ('onMessage the XML marshaled response: <?xml' in line) or \
                                    ('AdapterExecuter executeFramework Response <?xml' in line) or \
                                    ('PosSofBusines I SOF:PosSofBusinessImpl updateDebit  debitId:' in line) or \
                                    ('><Response>' in line):
                                messages.append(line)
                                print('thread ended and about to enter new thread. Processing Stopped')
                                break

                            if ('SOF:FawryReqResUtil generateErrorResponse' in line) or 'SOF:MerchantEJB process The sent response: MsgCode =' in line:
                                messages.append(line)
                                print('thread ended and about to enter new thread. Processing Stopped')
                                break

                            if search_type == 'archive':
                                messages.append(line)
                                for x in range(y + 1, y + 300):
                                    line = lines[x]
                                    if (adaptor_threads[l][0] in line) or re.match('^\[[0-9]+/[0-9]+/[0-9]+', line):
                                        break
                                    else:
                                        messages.append(lines[x])
                            elif search_type == 'archive_msgs':
                                print(adaptor_threads)
                                for error in error_list:
                                    if error.ErrorString in line:
                                        messages.append(line)
                                        for x in range(y + 1, y + 300):
                                            line = lines[x]
                                            if (adaptor_threads[l][0] in line) or re.match('^\[[0-9]+/[0-9]+/[0-9]+', line):
                                                break
                                            else:
                                                messages.append(lines[x])
                                        break

     return messages
    except Exception as e:
        print(str(e))
        traceback.print_exc()


def logs_error_logic(messages, result, component_type):
    if str(component_type) == 'Switch':
        for message in messages:
            if 'The Biller Outgoing Request is' in message:
                result.append('The request is sent to biller successfully')
                msg_index = messages.index(message)
                if len(messages) != 1:
                    logs_error_logic(messages[msg_index+1:], result, component_type)
                break
            elif 'The Biller Incomming Response is' in message:
                result.append('The Response is received from biller')
                msg_index = messages.index(message)
                response_statuses = SoapResponseStatus.objects.filter(name='PmtNotifyRs')
                for response_status in response_statuses:
                    if response_status.status in messages[msg_index+1]:
                        result.append(response_status.description)
                if len(messages) != 1:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'SwitchUtil S logException' in message:
                result.append('The request failed in SW duo to exception, root cause:')
                response_statuses = SoapResponseStatus.objects.filter(name='PmtNotifyRs')
                msg_index = messages.index(message)
                for msg in messages[msg_index+1:]:

                    for response_status in response_statuses:
                        if response_status.status in msg:
                            result.append(response_status.description)

                    if re.match('^\[[0-9]+/[0-9]+/[0-9]+', msg):
                        msg_index = messages.index(msg)
                        break
                result.append('Please contact support if issue cannot be detected')
                if len(messages) != 1:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif '###Finished Handling new thread with async:' in message:
                result.append("#####################")
    else:
        print(messages)
        for message in messages:
            if 'Gateway recieved message' in message or 'Gateway Received Message' in message:
                result.append('The request is received by GW from channel')
                msg_index = messages.index(message)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index+1:], result, component_type)
                break
            elif '- Message sent to Switch:' in message:
                result.append('The request is sent to SW successfully from GW')
                msg_index = messages.index(message)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index+1:], result, component_type)
                break
            elif '- Message received from Switch' in message:
                result.append('The Response is received by GW from SW ')
                msg_index = messages.index(message)
                response_statuses = SoapResponseStatus.objects.filter(name__in=['PmtAdvRs', 'PmtAddRs', 'BillInqRs'])
                for response_status in response_statuses:
                    if response_status.status in messages[msg_index+1]:
                        result.append(response_status.description)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'Gateway sent message' in message:
                result.append('The Response is sent to channel from GW')
                msg_index = messages.index(message)
                response_statuses = SoapResponseStatus.objects.filter(name__in=['PmtAddRs'])
                for response_status in response_statuses:
                    if response_status.status in messages[msg_index]:
                        result.append(response_status.description)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif '- Message sent to Wallet' in message:
                result.append('The request is sent to adaptor successfully from GW')
                msg_index = messages.index(message)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'executeService Message sent to SourceOfFund:' in message:
                result.append('The request is sent to sof successfully from GW')
                msg_index = messages.index(message)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'executeService Message received from SourceOfFund' in message:
                result.append('The Response is received by GW from SOF ')
                msg_index = messages.index(message)
                response_statuses = SoapResponseStatus.objects.filter(name__in=['DebitAddRs'])
                for response_status in response_statuses:
                    if response_status.status in messages[msg_index+1]:
                        result.append(response_status.description)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'executeService Message received from Wallet' in message:
                result.append('The Response is received by GW from adaptor')
                msg_index = messages.index(message)
                response_statuses = SoapResponseStatus.objects.filter(name__in=['DebitAddRs'])
                for response_status in response_statuses:
                    if response_status.status in messages[msg_index+1]:
                        result.append(response_status.description)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'AdapterExecuter executeFramework Request' in message:
                result.append('The Response is received by adaptor from GW')
                msg_index = messages.index(message)
                response_statuses = SoapResponseStatus.objects.filter(name__in=['DebitAddRs'])
                for response_status in response_statuses:
                    if response_status.status in messages[msg_index+1]:
                        result.append(response_status.description)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'AdapterExecuter executeFramework Response' in message:
                result.append('The Response is received by GW from adaptor')
                msg_index = messages.index(message)
                response_statuses = SoapResponseStatus.objects.filter(name__in=['DebitAddRs'])
                for response_status in response_statuses:
                    if response_status.status in messages[msg_index+1]:
                        result.append(response_status.description)
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif 'Exception' in message:
                result.append('The request failed duo to exception, root cause:')
                response_statuses = SoapResponseStatus.objects.filter(name__in=['PmtAddRs', 'DebitAddRs', 'BillInqRs', 'PmtAdvRs'])
                msg_index = messages.index(message)
                for msg in messages[msg_index+1:]:
                    for response_status in response_statuses:
                        if response_status.status in msg:
                            result.append(response_status.description)
                    if re.match('^\[[0-9]+/[0-9]+/[0-9]+', msg):
                        msg_index = messages.index(msg)
                        break
                result.append('Please contact support if issue cannot be detected')
                if len(messages) > 0:
                    logs_error_logic(messages[msg_index + 1:], result, component_type)
                break
            elif '###Finished Handling new thread with async:' in message:
                result.append("#####################")
    return


def soapTool(ip, port, group, out_queue, reqs, typeOrService):
    output = f'\n<b>{ip}</b>\n<b>=====================</b>\n'
    in_queue = queue.Queue()
    threads_list = list()
    for request in reqs:
        t = threading.Thread(target=sendSoapReq, args=(ip, port, group, in_queue, request))
        t.start()
        threads_list.append(t)

    for t in threads_list:
        t.join()

    while not in_queue.empty():
        if typeOrService == 'REQUEST':

            result = in_queue.get()

            output += result
            output += '\n'
        else:
            result = in_queue.get()
            out_queue.put([typeOrService.name, reqs[0].name, ip, result])

    if typeOrService == 'REQUEST':
        out_queue.put(output)

    return

def sendSoapReq(ip, port, group, in_queue, request, is_job_call=False):
    output = ''
    try:
        req_headers = SoapRequestHeaders.objects.filter(request_id__id=request.id)
        headers = {'content-type': request.content_type, }  # headers = {'content-type': 'application/soap+xml'}
        for header in req_headers:
            headers[header.key] = header.value

        url = "http://" + ip + ':' + port + request.uri
        body = request.body

        if request.type == 'post':
            response = requests.post(url, data=body.encode('UTF-8'), headers=headers, timeout=30)
        else:
            response = requests.get(url, headers=headers)

        response.raise_for_status()
    except requests.exceptions.ConnectionError as err:
        if is_job_call:
            addRequestMonitorRecord(request, group, ip, datetime.datetime.now(), str(err))

        print(err)
        output += f'<b>{request.name}</b>, url: {url}, Response: {err}\n----------------------------------------\n'
    except HTTPError as err:
        if is_job_call:
            addRequestMonitorRecord(request, group, ip, datetime.datetime.now(), str(err))

        print(err)
        output += f'<b>{request.name}</b>, url: {url}, Response: {err}\n----------------------------------------\n'
    except requests.exceptions.ReadTimeout as err:
        if is_job_call:
            addRequestMonitorRecord(request, group, ip, datetime.datetime.now(), str(err))

        print(err)
        output += f'<b>{request.name}</b>, url: {url}, Response: Timeout..\n----------------------------------------\n'
    else:

        if ';' in request.success_status:
            accepted_responses = request.success_status.split(';')
            # print(f'{request.name}, url: {url}, Response: success..')
            for accepted_response in accepted_responses:
                if accepted_response in response.content.decode("utf-8"):
                    output += f'<b>{request.name}</b>, url: {url}, Response: success..\n----------------------------------------\n'
                    break
        elif request.success_status in response.content.decode("utf-8"):
            #print(f'{request.name}, url: {url}, Response: success..')
            output += f'<b>{request.name}</b>, url: {url}, Response: success..\n----------------------------------------\n'
        elif '<StatusDesc>Sender is not authorized.</StatusDesc>' in response.content.decode("utf-8"):
            #print(f'{request.name}, url: {url}, Response: Sender not authorized.')
            output += f'<b>{request.name}</b>, url: {url}, Response: Sender not authorized.\n----------------------------------------\n'
        else:
            #print(f'{request.name}, url: {url}, Response: Business Exception.')
            output += f'<b>{request.name}</b>, url: {url}, Response: Business Exception.\n----------------------------------------\n'

    in_queue.put(output)
    return


def functionalToolsFn(toolName, groupname):
    print("Thread in functionalToolsFn: ")

    if toolName == "CheckAppWebSrv":
        if groupname != '':
            adaptors = AdaptorsAndApps.objects.filter(group_availability__groupName=groupname)

        else:
            adaptors = AdaptorsAndApps.objects.all()

        for adaptor in adaptors:
            response = requests.get(adaptor.endpoint)


def read_template(filename):
    """
    Returns a Template object comprising the contents of the
    file specified by filename.
    """

    with open("C:\MonitoringToolProject\\" + filename, 'r', encoding='utf-8') as template_file:
        template_file_content = template_file.read()
    return Template(template_file_content)


def form_table_entries(title, headers, results):
    table = '<div style="margin:10px"><h2>' + title + "</h2>\n" + "<table>\n<tr class= 'info-class'>\n"

    for x in range(len(headers)):
        cell = "<th>" + headers[x] + "</th>\n"
        table += cell

    table += "</tr>\n"

    for entry in results:
        table += "<tr class= 'info-class'>\n"
        row = ""
        for x in range(len(headers)):
            try:
                value = entry[x]
                if value in ('missing', 'exceed', 'low'):
                    cell = "<td class='thresholdalert'>" + str(value) + "</td>\n"
                else:
                    cell = "<td>" + str(value) + "</td>\n"
                row += cell
            except IndexError as e:
                cell = "<td class='dimmed'></td>\n"
                row += cell
            except TypeError as e:
                cell = "<td class='dimmed'></td>\n"
                row += cell
            except cx_Oracle.DatabaseError as e:
                cell = "<td class='dimmed'></td>\n"
                row += cell

        table += row
        table += "</tr>\n"
    table += "</table></div>"

    return table


def ach_checker():
    host = Server.objects.get(IP='10.100.44.161')
    srv_conn = ServerConnection(host)
    srv_conn.connect_ssh()
    if (True):
        command = 'cd /opt/IBM/home/dsadm/ETL_Extract/RTL_Batches/archive ; ls -ltr'
        temp = srv_conn.exec_command(command, True).splitlines()
        rec = temp[-1].replace('  ', ' ').replace('  ', ' ').split(' ')
        print(rec[5], rec[6], rec[7], rec[8], rec[9])

        #send_mail(rec[9], rec[5], 'group', 'curr_count', 'max', 'sev', 'moustafa.mamdouh@fawry.com', 'ach')


    if(True):
        command = 'cd /opt/IBM/home/dsadm/ETL_Extract/Accpt_Batches/Archive ; ls -ltr'
        temp = srv_conn.exec_command(command, True).splitlines()
        rec1 = temp[-1].replace('  ', ' ').replace('  ', ' ').split(' ')
        rec2 = temp[-2].replace('  ', ' ').replace('  ', ' ').split(' ')
        print(rec1[5], rec1[6], rec1[7], rec1[8], rec1[9])
        print(rec2[5], rec2[6], rec2[7], rec2[8], rec2[9])

        #send_mail(rec1[9], rec1[5], 'group', 'curr_count', 'max', 'sev', 'moustafa.mamdouh@fawry.com', 'ach')
        #send_mail(rec2[9], rec2[5], 'group', 'curr_count', 'max', 'sev', 'moustafa.mamdouh@fawry.com', 'ach')

    if (True):
        command = 'cd /opt/IBM/home/dsadm/ETL_Extract/Finance_Financial_Extract ; ls -ltr'
        temp = srv_conn.exec_command(command, True).splitlines()
        print(temp[-1])
        print(temp[-2])


@permission_required('auth.test_url', login_url="/login", raise_exception=True)
def test(request):

    #db_queries_checker()
    #result = was_queues_checker()
    #background_env_monitor()
    ACH_success_accounts_job_checker()
    #api_requests_checker()
    #ach_checker()
    #permissions()
    #abnormal_trx_monitor()
    #while ExceptionsMonitor.objects.count():
    #    ExceptionsMonitor.objects.all()[0].delete()
    return HttpResponse("Success")


def permissions():
    user = User.objects.get(username = 'admin')
    group = AdminGroup.objects.get(name='BusinessOperations')
    permissions = Permission.objects.filter(group=group)
    print(permissions)
    #content_type = ContentType.objects.get_for_model(User)

    #permission = Permission.objects.create(
    #    codename='can_monitor_apis',
    #    name='can_monitor_apis',
    #    content_type=content_type,
    #)
